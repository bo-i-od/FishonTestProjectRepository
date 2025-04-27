import json
import os
from pathlib import Path

from openpyxl import *


from common.error import PluralElementError, FindNoElementError
from tools import baseDataRead
from tools.decl2py import *


class ExcelTools:
    def __init__(self, root_path):
        self.excel_path = root_path
        self.base_data_path = self.excel_path.split("策划模板导出工具/")[0] + r"ElementData/BaseData/"
        file_path = os.path.join(os.path.dirname(__file__))
        # print(file_path)
        # # 获取当前工作目录
        # current_dir = os.getcwd()
        # 获取父目录
        self.root_dir = os.path.abspath(os.path.dirname(file_path))


    def get_worksheet(self, book_name, sheet_name):
        path = self.excel_path + book_name
        workbook = load_workbook(path, data_only=True)
        worksheet = workbook[sheet_name]
        return worksheet

    @staticmethod
    def get_column_index(worksheet, header):
        column_index_res = 0
        for column_index_cur in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(4, column_index_cur).value
            if cell_value == header:
                column_index_res = column_index_cur
                break
        return column_index_res

    @staticmethod
    def get_row_index(worksheet, column_index, target_value):
        row_index_res = 0
        for row_index_cur in range(1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row_index_cur, column_index).value
            if cell_value == target_value:
                row_index_res = row_index_cur
                break
        return row_index_res


    @staticmethod
    def get_value_from_key(table_data, header_key, header_value, key):
        index = table_data[header_key].index(key)
        value = int(table_data[header_value][index])
        return value


    def same_row_different_column_convert(self, worksheet, source_header, target_header, source):
        try:
            column_index_source = self.get_column_index(worksheet, source_header)
            column_index_target = self.get_column_index(worksheet, target_header)
            row_index = self.get_row_index(worksheet, column_index_source, source)
            return worksheet.cell(row_index, column_index_target).value
        except:
            return None

    def same_row_different_column_convert_list(self, worksheet, source_header, target_header, source_list):
        res_list = []
        column_index_source = self.get_column_index(worksheet, source_header)
        column_index_target = self.get_column_index(worksheet, target_header)
        for source in source_list:
            row_index = self.get_row_index(worksheet, column_index_source, source)
            res_list.append(worksheet.cell(row_index, column_index_target).value)
        return res_list


    def get_table_struct(self, book_name):
        res_dict = {}
        worksheet = self.get_worksheet(book_name=book_name, sheet_name="OrgData")
        ItemName1_column_index = 3
        cell_value_pre = ""
        row_index_cur = 2
        while worksheet.cell(row_index_cur, ItemName1_column_index - 1).value is not None:
            cell_value = worksheet.cell(row_index_cur, ItemName1_column_index).value
            # 数组的添加索引列表
            if cell_value == cell_value_pre:
                res_dict[cell_value].append(row_index_cur - 1)
                row_index_cur += 1
                continue
            # 普通的也直接添加
            if cell_value is not None:
                res_dict[cell_value] = [row_index_cur - 1]
                cell_value_pre = cell_value
                row_index_cur += 1
                continue
            dict_list_len = len(res_dict[cell_value_pre])
            res_dict[cell_value_pre][dict_list_len - 1] = {}
            ItemName2_column_index = ItemName1_column_index + 1
            res_dict[cell_value_pre][dict_list_len - 1][worksheet.cell(row_index_cur - 1, ItemName2_column_index).value] = [row_index_cur - 2]
            cur = row_index_cur
            while worksheet.cell(cur, ItemName1_column_index).value is None and worksheet.cell(cur, ItemName1_column_index - 1).value is not None:
                res_dict[cell_value_pre][dict_list_len - 1][worksheet.cell(cur, ItemName2_column_index).value] = [cur - 1]
                cur += 1
            row_index_cur = cur
        return res_dict, row_index_cur - 2

    def get_table_data(self, book_name):
        table_data, table_data_len = self.get_table_struct(book_name=book_name)
        worksheet = self.get_worksheet(book_name=book_name, sheet_name="模板数据")
        # 去除空白列
        bias_list = self.get_bias_list(worksheet, table_data_len)
        # table_data = {}
        for key in table_data:
            if len(table_data[key]) == 1:
                if isinstance(table_data[key][0], int):
                    table_data[key] = self.get_column_data(worksheet, table_data[key][0], bias_list)
                    continue
                # table_data[key] = []
                # table_data[key].append({})
                for key_sub in table_data[key][0]:
                    table_data[key][0][key_sub] = self.get_column_data(worksheet, table_data[key][0][key_sub][0], bias_list)
                continue
            cur = 0
            while cur < len(table_data[key]):
                if isinstance(table_data[key][cur], int):
                    table_data[key][cur] = self.get_column_data(worksheet, table_data[key][cur], bias_list)
                    cur += 1
                    continue
                for key_sub in table_data[key][cur]:
                    table_data[key][cur][key_sub] = self.get_column_data(worksheet, table_data[key][cur][key_sub][0], bias_list)
                cur += 1
        return table_data

    def get_table_data_object_list_by_excel(self, book_name):
        table_data_object_list = []
        table_data, table_data_len = self.get_table_struct(book_name=book_name)
        worksheet = self.get_worksheet(book_name=book_name, sheet_name="模板数据")
        cur = 6
        while worksheet.cell(cur, 1).value is not None:
            table_data_template = table_data.copy()
            row_data = self.get_row_data(worksheet, row_index=cur, table_data_len=table_data_len)
            table_data_object = self.fill_template(template=table_data_template, data=row_data)
            table_data_object_list.append(table_data_object)
            cur += 1
        return table_data_object_list

    def get_table_data_detail(self, book_name):
        prefix = book_name.split('.')[0]
        table_data_detail = baseDataRead.convert_to_json(path=self.base_data_path, prefix=prefix)
        return table_data_detail

    def get_table_data_object_list(self, book_name):
        table_data_object_list, _, _ = self.get_table_data_detail(book_name=book_name)
        return table_data_object_list


    def get_table_data_object_list_by_key_value(self, key, value, book_name=None, table_data_detail=None):
        res = []
        if table_data_detail is None:
            table_data_detail = self.get_table_data_detail(book_name=book_name)
        table_data_object_list, structs, prefix = table_data_detail

        # 给值转为正确的类型
        if prefix.upper() not in structs:
            return res
        type_value = structs[prefix.upper()][key][0]
        if type_value == "int":
            value = int(value)
        elif type_value == "float":
            value = float(value)
        elif type_value == "string":
            value = str(value)
        for table_data_object in table_data_object_list:
            if key not in table_data_object:
                continue
            if table_data_object[key] != value:
                continue
            res.append(table_data_object)
        return res

    def get_table_data_object_by_key_value(self, key, value, book_name=None, table_data_detail=None):
        if table_data_detail is None:
            table_data_detail = self.get_table_data_detail(book_name=book_name)
        table_data_object_list, structs, prefix = table_data_detail
        type_value = structs[prefix.upper()][key][0]
        if type_value == "int":
            value = int(value)
        elif type_value == "float":
            value = float(value)
        elif type_value == "string":
            value = str(value)
        res = None
        cur = 0
        for table_data_object in table_data_object_list:
            if key not in table_data_object:
                continue
            if table_data_object[key] != value:
                continue
            cur += 1
            res = table_data_object
        if cur != 1:
            raise PluralElementError("表中键值的匹配项不是唯一，请使用get_table_data_object_list_by_key_value")
        return res


    @staticmethod
    def fill_template(template, data):
        result = str(template)
        for i, value in enumerate(data, start=1):
            if isinstance(value, str):
                result = result.replace(f'[{i}]', f"'{value}'")
            else:
                result = result.replace(f'[{i}]', str(value))
        result = result.replace("'", '"')
        result = json.loads(result)
        return result

    @staticmethod
    def get_row_data(worksheet, row_index, table_data_len):
        row_data = []
        # 第六行开始
        cur = 1
        while cur <= table_data_len:
            row_data.append(worksheet.cell(row_index, cur).value)
            cur += 1
        return row_data


    @staticmethod
    def get_bias_list(worksheet, table_data_len):
        bias_list = []
        bias = 0
        cur = 1
        while cur < table_data_len + 1:
            if worksheet.cell(2, cur + bias).value is None:
                bias += 1
                continue
            bias_list.append(bias)
            cur += 1
        return bias_list

    @staticmethod
    def get_column_data(worksheet, column_index, bias_list):
        column_data = []
        # 第六行开始
        cur = 6
        while worksheet.cell(cur, 1).value is not None:
            index = column_index - 1
            column_data.append(worksheet.cell(cur, column_index + bias_list[index]).value)
            cur += 1
        return column_data



from activities.decl.TIMER_MAIN import TIMER_MAIN
class ExcelToolsForActivities(ExcelTools):
    def __init__(self, root_path):
        super().__init__(root_path)

    def write_data_txt(self, name: str, blocks: str = None, json_object_list=None, instance_object_list=None):
        """
            blocks, json_object_list, instance_object_list三选一保存到{name}.data.txt中
        """
        if instance_object_list:
            instance_list_to_json_list(instance_object_list=instance_object_list)
        if json_object_list:
            blocks = json_list_to_blocks(json_object_list=json_object_list, name=name.lower())
        Path(self.base_data_path + name + ".data.txt").write_text("\n" + blocks, encoding="utf-16")

    def add_object(self, key: str=None, value=None, book_name: str = None, table_data_detail=None, json_object: dict = None, instance_object=None):
        """
            增加object到txt中
            key: 特征键 value：特征值 可不写，写了如果有重复项就不添加
            book_name，table_data_detail二选一
            json_object，instance_object二选一
        """
        if not table_data_detail:
            table_data_detail = self.get_table_data_detail(book_name=book_name)
        table_data_object_list, structs, prefix = table_data_detail
        if key and value:
            # 已经有该键值对的数据
            if self.get_table_data_object_list_by_key_value(key=key, value=value, table_data_detail=table_data_detail):
                print(f"{prefix}表中已经有{key}:{value}对应的数据，故跳过新增")
                return False
            if instance_object:
                json_object = instance_to_json(instance_object=instance_object)
            table_data_object_list.append(json_object)
            self.write_data_txt(name=prefix, json_object_list=table_data_object_list)
            return True

        if instance_object:
            json_object = instance_to_json(instance_object=instance_object)
        table_data_object_list.append(json_object)
        self.write_data_txt(name=prefix, json_object_list=table_data_object_list)
        return True

    def add_objects(self, key: str=None, value_list=None, book_name: str = None, table_data_detail=None, json_object_list: list = None, instance_object_list: list =None):
        """
            批量增加object到txt中
            key: 特征键 value：特征值 可不写，写了如果有重复项就不添加
            book_name，table_data_detail二选一
            json_object，instance_object二选一
        """
        if not table_data_detail:
            table_data_detail = self.get_table_data_detail(book_name=book_name)

        cur = 0
        while cur < len(value_list):
            if json_object_list:
                self.add_object(key=key, value=value_list[cur], table_data_detail=table_data_detail, json_object=json_object_list[cur])
                cur += 1
                continue
            self.add_object(key=key, value=value_list[cur], table_data_detail=table_data_detail, instance_object=instance_object_list[cur])
            cur += 1

    def remove_object(self, key: str, value, book_name: str = None, table_data_detail=None, json_object: dict = None, instance_object=None):
        """
            移除所有指定键值的object
            key: 特征键
            value：特征值
            book_name，table_data_detail二选一
            json_object，instance_object二选一
        """
        if not table_data_detail:
            table_data_detail = self.get_table_data_detail(book_name=book_name)
        table_data_object_list, structs, prefix = table_data_detail
        cur = 0
        while cur < len(table_data_object_list):
            table_data_object = table_data_object_list[cur]
            if key not in table_data_object:
                cur += 1
                continue
            if value != table_data_object[key]:
                cur += 1
                continue
            if instance_object:
                json_object = instance_to_json(instance_object=instance_object)
            table_data_object_list.remove(json_object)
            cur += 1
        self.write_data_txt(name=prefix, json_object_list=table_data_object_list)

    def remove_objects(self, key: str, value_list, book_name: str = None, table_data_detail=None, json_object_list:list = None, instance_object_list: list=None):
        """
            移除所有指定键值的object
            key: 特征键
            value_list：特征值列表
            book_name，table_data_detail二选一
            json_object，instance_object二选一
        """
        if not table_data_detail:
            table_data_detail = self.get_table_data_detail(book_name=book_name)
        cur = 0
        while cur < len(value_list):
            if json_object_list:
                self.remove_object(key=key, value=value_list[cur], table_data_detail=table_data_detail, json_object=json_object_list[cur])
                cur += 1
                continue
            self.remove_object(key=key, value=value_list[cur], table_data_detail=table_data_detail, instance_object=instance_object_list[cur])
            cur += 1

    def change_object(self, key: str, value, book_name: str = None, table_data_detail=None, json_object: dict = None, instance_object=None):
        """
            改变所有指定键值的object
            key: 特征键
            value：特征值
            book_name，table_data_detail二选一
            json_object，instance_object二选一
        """
        if not table_data_detail:
            table_data_detail = self.get_table_data_detail(book_name=book_name)
        table_data_object_list, structs, prefix = table_data_detail
        # print(table_data_object_list)
        cur = 0
        while cur < len(table_data_object_list):
            table_data_object = table_data_object_list[cur]
            if key not in table_data_object:
                cur += 1
                continue
            if value != table_data_object[key]:
                cur += 1
                continue
            if instance_object:
                json_object = instance_to_json(instance_object=instance_object)
            table_data_object_list[cur] = json_object
            cur += 1
        self.write_data_txt(name=prefix, json_object_list=table_data_object_list)

    def change_objects(self, key: str, value_list, book_name: str = None, table_data_detail=None, json_object_list: list = None, instance_object_list: list=None, is_plural=False):
        """
            改变所有指定键值的object
            key: 特征键
            value_list：特征值列表
            book_name，table_data_detail二选一
            json_object，instance_object二选一
        """
        if not table_data_detail:
            table_data_detail = self.get_table_data_detail(book_name=book_name)
        cur = 0
        while cur < len(value_list):
            if json_object_list:
                self.change_object(key=key, value=value_list[cur], table_data_detail=table_data_detail, json_object=json_object_list[cur])
                cur += 1
                continue
            self.change_object(key=key, value=value_list[cur], table_data_detail=table_data_detail, instance_object=instance_object_list[cur])
            cur += 1


    def get_object(self,  key: str, value, book_name: str = None, table_data_detail=None, cls: type = None, is_plural=False):
        """
            获取json_object和instance_object
            key: 特征键
            value：特征值
            book_name，table_data_detail二选一
            json_object，instance_object二选一
            cls：instance_object的类型 当不写时instance_object的值为None
        """
        if not table_data_detail:
            table_data_detail = self.get_table_data_detail(book_name=book_name)
        json_object_list = self.get_table_data_object_list_by_key_value(key=key, value=value, book_name=book_name, table_data_detail=table_data_detail)
        if not json_object_list:
            return None, None
        json_object = json_object_list[0]
        instance_object_list = None
        instance_object = None
        if cls:
            instance_object_list = json_list_to_instance_list(json_object_list=json_object_list, cls=cls)
            instance_object = instance_object_list[0]
        if is_plural:
            return json_object_list, instance_object_list
        return json_object, instance_object

    def get_objects(self, key, value_list, book_name: str = None, table_data_detail=None, cls: type = None):
        """
            获取json_object_list, instance_object_list
            需要cls=instance_object的类型，否则instance_object_list=[]
        """
        if not table_data_detail:
            table_data_detail = self.get_table_data_detail(book_name=book_name)
        json_object_list = []
        instance_object_list = []
        cur = 0
        while cur < len(value_list):
            json_object, instance_object = self.get_object(key=key, value=value_list[cur], table_data_detail=table_data_detail, cls=cls)
            json_object_list.append(json_object)
            instance_object_list.append(instance_object)
            cur += 1
        return json_object_list, instance_object_list


    def timer_main(self, timer_id:int, time_start: str, time_end: str, timer_main_detail=None):
        """
            更改指定timer_id的openTime和endTime
        """
        if timer_main_detail is None:
            timer_main_detail = self.get_table_data_detail(book_name="TIMER_MAIN.xlsm")
        instance_object: TIMER_MAIN
        json_object, instance_object = self.get_object(key="timerID", value=timer_id, table_data_detail=timer_main_detail, cls=TIMER_MAIN)
        instance_object.openTime = time_start
        instance_object.endTime = time_end
        print(instance_object)
        self.change_object(key="timerID", value=timer_id, instance_object=instance_object, table_data_detail=timer_main_detail)


    def fish_bag_id_to_detail(self, fish_bag_id, table_object_detail=None):
        """
            name              鱼卡包名称
            itemTpId          fish_bag_id
            fishCardCount     一包里鱼卡总数
            fishBagFishery    1-4是tier1-4卡包  x003xx是x003xx卡包
            fishBagType       1是普通卡包 2是hidden卡包 3是boss卡包
        """
        if table_object_detail is None:
             table_object_detail = baseDataRead.convert_to_json(path=self.root_dir + "/activities/customTables/", prefix="FISH_BAG")
        json_object_list, _, _ = table_object_detail
        for json_object in json_object_list:
            if json_object["itemTpId"] != fish_bag_id:
                continue
            return json_object
        return None

    def change_fish_bag_fishery(self, fish_bag_id, fishery_id, table_object_detail=None):
        """
            将fish_bag_id转换成fishery_id对应的卡包id，如果不存在返回None
        """
        if table_object_detail is None:
            table_object_detail = baseDataRead.convert_to_json(path=self.root_dir + "/activities/customTables/", prefix="FISH_BAG")
        detail = self.fish_bag_id_to_detail(fish_bag_id=fish_bag_id, table_object_detail=table_object_detail)
        if detail is None:
            return None
        if detail["fishBagFishery"] == 5:
            return fish_bag_id
        return self.get_fish_bag(fishery_id=fishery_id, fish_bag_type=detail["fishBagType"],  fish_card_count=detail["fishCardCount"], table_object_detail=table_object_detail)


    def get_fish_bag(self, fishery_id=None, fish_bag_type=None, fish_card_count=None, table_object_detail=None):
        """
            fishBagType 1是普通卡包 2是hidden卡包 3是boss卡包
            fishery_id  1-4是tier1-4卡包  x003xx是x003xx卡包
            fish_card_count  一包里鱼卡总数
            table_object_detail是FISH_BAG.data.txt的数据
        """
        if table_object_detail is None:
            table_object_detail = baseDataRead.convert_to_json(path=self.root_dir + "/activities/customTables/", prefix="FISH_BAG")
        json_object_list, _, _ = table_object_detail
        for json_object in json_object_list:
            if fishery_id and "fishBagFishery" not in json_object:
                continue
            if fish_bag_type and "fishBagType" not in json_object:
                continue
            if fish_card_count and "fishCardCount" not in json_object:
                continue
            if fishery_id and json_object["fishBagFishery"] != fishery_id:
                continue
            if fish_bag_type and json_object["fishBagType"] != fish_bag_type:
                continue
            if fish_card_count and json_object["fishCardCount"] != fish_card_count:
                continue
            return json_object["itemTpId"]
        raise FindNoElementError("没有找到对应的鱼卡包")

    def group_id_to_timer_id(self, group_id):
        """
            MISSION_GROUP中根据group_id查timer_id
        """
        table_data_object = self.get_table_data_object_by_key_value(book_name="MISSION_GROUP.xlsm", key="groupId", value=group_id)
        timer_id = table_data_object["openArg"]
        return timer_id

    def get_max_value(self, key, table_object_detail):
        """
            获取表中该键最大的值
        """
        json_object_list, _ , _ = table_object_detail
        max_value = 0
        for json_object in json_object_list:
            if key not in json_object:
                # print(json_object)
                continue
            if max_value > json_object[key]:
                continue
            if json_object[key] == 9999999:
                continue
            if json_object[key] == 99999999:
                continue
            if json_object[key] == 9999999:
                continue
            max_value = json_object[key]
        return max_value

    def get_min_value_more_than_start(self, key, table_object_detail, start, long=1):
        """
            获得大于一个指定数能容纳连续long个数的最小数
        """
        json_object_list, _, _ = table_object_detail
        existing_values = {obj[key] for obj in json_object_list if key in obj}

        n = start
        while True:
            # 检查从n开始的long个连续数是否均未被使用
            if all((n + i) not in existing_values for i in range(long)):
                return n
            n += 1


    def get_fish_id_list(self, fishery_id, fisheries_detail=None):
        """函数功能简述
            根据fishery_id获取该渔场的鱼id列表

        参数:
            fishery_id: 渔场id
        """
        if fisheries_detail is None:
            fisheries_detail = self.get_table_data_detail(book_name="FISHERIES.xlsm")
        table_data_object = self.get_table_data_object_by_key_value(key="tpId", value=fishery_id, table_data_detail=fisheries_detail)
        fish_list = table_data_object["fish"]
        activity_fish_list = []
        if "activityFishNotShow" in table_data_object:
            activity_fish_list = table_data_object["activityFishNotShow"]
        res_list = []
        for fish in fish_list:
            if not fish:
                continue
            if activity_fish_list and fish in activity_fish_list:
                continue
            res_list.append(fish)
        return res_list

    def get_fish_type(self, fish_id, fish_detail=None):
        """
            查FISH表里的fishType
        """
        if fish_detail is None:
            fish_detail = self.get_table_data_detail(book_name="FISH.xlsm")
        table_data_object = self.get_table_data_object_by_key_value(key="tpId", value=fish_id, table_data_detail=fish_detail)
        return table_data_object["fishType"]

    def get_fish_class(self, fish_id, fish_detail=None):
        """
            查FISH表里的fishClass
        """
        if fish_detail is None:
            fish_detail = self.get_table_data_detail(book_name="FISH.xlsm")
        table_data_object = self.get_table_data_object_by_key_value(key="tpId", value=fish_id, table_data_detail=fish_detail)
        return table_data_object["fishClass"]

    def get_rod(self, fishery_id, rarity, fisheries_detail=None, fishing_rod_detail=None):
        """
            根据渔场id和稀有度查对应的鱼竿
            rarity 2蓝 3紫 4黄
        """
        if fisheries_detail is None:
            fisheries_detail = self.get_table_data_detail(book_name="FISHERIES.xlsm")
        if fishing_rod_detail is None:
            fishing_rod_detail = self.get_table_data_detail(book_name="FISHING_ROD.xlsm")
        fishery_detail = self.get_fishery_detail(fishery_id=fishery_id, fisheries_detail=fisheries_detail)
        fishery_rank = fishery_detail["fisheriesRank"]
        fishery_living = fishery_detail["fisheriesLiving"]
        table_data_object_list = self.get_table_data_object_list_by_key_value(key="rarity", value=rarity, table_data_detail=fishing_rod_detail)
        rod_list = []
        for table_data_object in table_data_object_list:
            if table_data_object["fisheriesRank"] != fishery_rank:
                continue
            if table_data_object["isFreshWater"] != fishery_living:
                continue
            rod_list.append(table_data_object["tpId"])
        return rod_list

    def get_rod_icon(self, rod_id, fishing_rod_detail=None):
        """
            根据rod_id查FISHING_ROD中displayicon
        """
        if fishing_rod_detail is None:
            fishing_rod_detail = self.get_table_data_detail(book_name="FISHING_ROD.xlsm")
        rod_icon = self.get_table_data_object_by_key_value(key="tpId", value=rod_id, table_data_detail=fishing_rod_detail)["displayicon"]
        return rod_icon

    def get_fishery_detail(self, fishery_id, fisheries_detail=None):
        """
            获取指定渔场json格式数据
        """
        if fisheries_detail is None:
            fisheries_detail = self.get_table_data_detail(book_name="FISHERIES.xlsm")
        table_data_object_list = self.get_table_data_object_list_by_key_value(key="tpId", value=fishery_id, table_data_detail=fisheries_detail)
        if not table_data_object_list:
            return None
        table_data_object = table_data_object_list[0]
        return table_data_object

    def get_fishery_fish_type_detail(self, fishery_id, fish_detail=None):
        """
            获取指定渔场各体型鱼的分布数据
            例如{"small":4, "medium": 4, "large": 4, "hidden": 2, "boss": 4, "rare": 3, "elite": 3, "monster": 6, "total": 25, "total_common": 15}
        """
        if fish_detail is None:
            fish_detail = self.get_table_data_detail(book_name="FISH.xlsm")
        fish_id_list = self.get_fish_id_list(fishery_id=fishery_id)
        fish_type_detail = {"small":0, "medium": 0, "large": 0, "hidden": 0, "boss": 0, "rare": 0, "elite": 0, "monster": 0, "total": len(fish_id_list), "total_common": 0}
        cur = 0
        while cur < len(fish_id_list):
            fish_id = fish_id_list[cur]
            fish_class = self.get_fish_class(fish_id=fish_id, fish_detail=fish_detail)
            if fish_class == 2:
                fish_type_detail["rare"] += 1
                cur += 1
                continue
            if fish_class == 3:
                fish_type_detail["elite"] += 1
                cur += 1
                continue
            if fish_class == 4:
                fish_type_detail["monster"] += 1
                cur += 1
                continue
            fish_type_detail["total_common"] += 1
            fish_type = self.get_fish_type(fish_id=fish_id, fish_detail=fish_detail)
            if fish_type == 1:
                fish_type_detail["small"] += 1
                cur += 1
                continue
            if fish_type == 2:
                fish_type_detail["medium"] += 1
                cur += 1
                continue
            if fish_type == 3:
                fish_type_detail["large"] += 1
                cur += 1
                continue
            if fish_type == 4:
                fish_type_detail["hidden"] += 1
                cur += 1
                continue
            if fish_type == 5:
                fish_type_detail["boss"] += 1
                cur += 1
                continue
            cur += 1
        return fish_type_detail

    def get_fishery_name(self, fishery_id, fisheries_language_detail=None):
        """
            根据fishery_id查FISHERIES_LANGUAGE中t_name
        """
        if fisheries_language_detail is None:
            fisheries_language_detail = self.get_table_data_detail(book_name="FISHERIES_LANGUAGE.xlsm")
        table_data_object = self.get_table_data_object_by_key_value(key="tpId", value=fishery_id, table_data_detail=fisheries_language_detail)
        return table_data_object["t_name"]

    def get_fish_name(self, fish_id, fish_language_detail=None):
        """
            根据fish_id查FISH_LANGUAGE中t_fishName
        """
        if fish_language_detail is None:
            fish_language_detail = self.get_table_data_detail(book_name="FISH_LANGUAGE.xlsm")
        table_data_object = self.get_table_data_object_by_key_value(key="tpId", value=fish_id, table_data_detail=fish_language_detail)
        return table_data_object["t_fishName"]

    def get_flash_card_id(self, fish_id, collection_base_detail=None):
        """
            根据fish_id查COLLECTION_BASE中闪卡id
        """
        if collection_base_detail is None:
            collection_base_detail = self.get_table_data_detail(book_name="COLLECTION_BASE.xlsm")
        return self.get_table_data_object_by_key_value(key="fishId", value=fish_id, table_data_detail=collection_base_detail)["collectionId"]







if __name__ == '__main__':
    et = ExcelTools("C:/trunkCHS/datapool/策划模板导出工具/")










