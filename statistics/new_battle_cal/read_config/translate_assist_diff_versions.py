#!/usr/bin/env python
# -*- coding:utf-8 -*-
# 功能：收集母表中的英文和中文，
# 英文作为国际化的参照语言直接使用，
# 而内容为中文时，表示直接使用的中文翻译结果，需要先翻译为英文，再从英文翻译到其它语言
#

from enum import Enum
from functools import cmp_to_key
import getopt
import logging
import pickle
import re
import shutil
import sys
import os
import codecs
import locale
import time

import openpyxl
# import xlsxwriter

class ClassifyFile(Enum):
    # 无表
    NoTables = 0
    # Localization 表
    LangTables = 1
    # Normal 表（非多语言表）
    NormalTables = 2
    # 全量表
    AllTables = 3

class Op(Enum):
    NoOp = 1
    # 仅从母表取,适用于首次收集时,认为中文文本全部无效
    Collect = 2
    # 全量收集，即从母表+其它已翻译的多语言数据表读取
    CollectAll = 3
    # 回写
    WriteBack = 4
    # 生成差分表
    GenerateDiffPart = 5
    # 生成2个版本的对比差异结果
    Generate2VersionDiff = 6
    # 生成3个版本的对比差分结果
    Generate3VersionDiff = 7
    # 把中文翻译结果合并到En母表，这是为国内版本支持的一个工具项
    MergeChsToEnBase = 8
    # 检查翻译结果中%s %d 格式化翻译前后顺序不一致的情况
    CheckFormatSerial = 9
    # 把所有多语言txt中双引号内的双引号全部替换为指定字符
    DoubleQuote = 10
    # 收集所有配置的id序列，得到一个基础表
    CollectAllTableId = 11
    # 修复整个配置表中多出的重复id，自动生成新值，并将最大值回写进
    FixDuplicateTableId = 12
    # 多语言id字段补齐完整
    LangPartFillToFull = 13
    # 扫描所有翻译内容，判定翻译结果中的格式化字符串是否正确
    ScanTranslate = 14
    # 修正多语言表的 id字段为 templateID
    FixIdToTemplateId = 15

g_op_enum_map = {enum.name.lower(): enum for enum in Op}

# 操作
g_op: Op = Op.Collect
g_op_str = g_op.name.lower()

g_op_files = ""

# 双引号中的双引号替换为单引号 '
g_double_quote_replace_char = '\''

# 脚本工具目录
g_china_712_path = ""
# 510版本原始表格内容目录
g_old_version_path = ""
# 810版本原始表格内容目录
g_new_version_path = ""

# 普通表模式，默认 False 即默认多语言表的处理模式
g_no_lang_mode = True

g_id_key = "id"
g_table_name_key = "tb"
# 纯 name 字符串
g_static_var_key = "__st_v"
g_name_key = "name"
g_table_data_name_key = "__tb_data"

g_part_key = "part"
g_value_key = "value"

g_dup_key = "dup"

g_contain_tables_key = "contain_tables"
g_contain_langs_key = "contain_langs"


##################
g_decl_struct_key = "__decl_struct"
g_decl_var_order_key = "__var_order"
g_analyse_var_type_key = "v_type"
g_analyse_var_name_key = "v_name"
g_analyse_var_if_array_key = "v_if_array"
g_analyse_var_if_custom_key = "v_if_custom"
g_analyse_var_issue_key = "v_issue"
g_analyse_var_index_key = "v_index"
# 是否是纯字符串，翻译表的模式下，只取纯字符串
g_analyse_var_is_pure_str_key = "v_is_pure_str"
# 是否忽略字段,name时为true
g_analyse_var_ignore_key = "v_ignore"
##################

g_sort_key = "sort_key"


g_en_key = "en"
g_chs_key = "chs"
g_content_key = "content"

g_global_810_key = "global810"
g_old_510_key = "old510"
g_china_712_key = "china712"

g_old_key = "old"
g_new_key = "new"

g_count_key = "count"
g_list_key = "list"
g_serial_key = "serial"
g_all_same_key = "all_same"

# 读取时得到的多语言列表
g_read_langs = {g_en_key:g_en_key,g_chs_key:g_chs_key}

g_dup_map_from_origin = {}

# 不翻译，直接作为结果的
g_const_strs = {"common_language":{g_id_key:{
    "1739", #  简体中文  # 简体中文
    "1740", #  繁体中文  # 繁体中文
    "1741", #  英语  # English
    "1815", #  越南语  # Tiếng Việt
    "1816", #  泰语  # ภาษาไทย
    "1817", #  土耳其语  # Türk dili
    "1818", #  德语  # Deutsch
    "1819", #  韩语  # 한국어
    "1880", #  波兰语  # Polskie
    "1881", #  葡萄牙语  # Português
    "1882", #  意大利语  # Italiano
    "1883", #  印尼语  # Bahasa Indonesia
    "1884", #  日语  # 日本語
    "1925", #  阿拉伯语  # "عربي"
    "1926", #  法语  # Français
    "1927", #  西班牙语  # Español
    "1928", #  俄语  # русский
    },"const_key":["t_panellanguage"]}
}

g_code_chinese = {
  # 需要自定义的对应关系
  "chs": "简体中文",
  "cht": "繁体中文",
  "rus": "俄语",

  # gpt给出的列表
  "zh_cn": "简体中文",
  "zh_tw": "繁体中文",
  "en": "英语",
  "vi": "越南语",
  "th": "泰语",
  "tr": "土耳其语",
  "de": "德语",
  "ko": "韩语",
  "pl": "波兰语",
  "pt": "葡萄牙语",
  "it": "意大利语",
  "id": "印尼语",
  "ja": "日语",
  "ar": "阿拉伯语",
  "fr": "法语",
  "esp": "西班牙语",
  "ru": "俄语",

}

g_test_tables = ["ITEM_MAIN_LANGUAGE.data.txt"]
g_test_tables = None

##################################################################
# 创建一个 logger 对象
logger = logging.getLogger('opt_logger')
logger.setLevel(logging.DEBUG)
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.DEBUG)
logger.addHandler(console_handler)
all_need_collector_log = []
log_near_char_count = 20
##################################################################
def get_file_extension(file_path):
    _, file_extension = os.path.splitext(file_path)
    return file_extension
#########################################
def get_file_name_no_ext(file_path):
    file_name, _ = os.path.splitext(file_path)
    return file_name
#########################################
def writeFile(file_path,str, encoding='utf-16'):
    file = open(file_path,'w', encoding=encoding)
    if file :
        file.write(str)
        file.close()
    else:
        print("error file %s open for write failed." % (file_path))
#########################################
def writeBinaryFile(file_path,str):
    file = open(file_path,'wb')
    if file :
        file.write(str)
        file.close()
    else:
        print("error file %s open for write failed." % (file_path))
#########################################
def readFile(file_path,encoding='utf-16'):
    if os.path.exists(file_path):
        file = open(file_path,'r', encoding=encoding)
        if file :
            content = file.read()
            file.close()
            return content
        else:
            print("error file %s open failed." % (file_path))
        pass
    else:
        print("error file %s not exists." % (file_path))
    return ""
def readBinaryFile(file_path):
    if os.path.exists(file_path):
        file = open(file_path,'rb')
        if file :
            content = file.read()
            file.close()
            return content
        else:
            print("error file %s open failed." % (file_path))
        pass
    else:
        print("error file %s not exists." % (file_path))
    return ""
#########################################
def readFileLines(file_path, encoding='utf-16'):
    if os.path.exists(file_path):
        file = open(file_path,'r', encoding=encoding)
        if file :
            content = file.readlines()
            file.close()
            return content
        else:
            print("error file %s open failed." % (file_path))
        pass
    else:
        print("error file %s not exists." % (file_path))
    return []
#########################################
def copyFile(src_file: str, dest_file: str):
    if os.path.exists(src_file):
        shutil.copy(src_file,dest_file)
#########################################
def copyTree(src_folder: str, dest_folder: str):
    if os.path.exists(src_folder):
        shutil.copytree(src_folder,dest_folder) # We inform you that server maintenance will occur at <color=#e5bf59> %s\n(UTC+8)</i></color>. During this time, you might experience a login interruption lasting up to <color=#e5bf59><i>%s</i></color> hours. As a thank you for your patience, we'll send you <color=#e5bf59>Diamonds x%s </color> after the maintenance. \nThank you for being so understanding!\n

#########################################
def removeTree(dest_folder: str):
    if os.path.exists(dest_folder):
        shutil.rmtree(dest_folder)
##################################################################
def _prechar_is_rdline(pre_char):
    return pre_char == '\\'
_find_assist_match_pairs = {'{':'}','"':'"',"'":"'"}
_find_assist_enter_stack_pairs = {'{':{'{':True,'"':True,"'":True},'"':{'{':False,"'":False},"'":{'{':False,'"':False}}
def find_right_f_edge(string):
    b_f_stack = []
    cur_f_char = "{"
    pre_char = None
    for i, char in enumerate(string):
        if not _prechar_is_rdline(pre_char) and char == _find_assist_match_pairs[cur_f_char]:
            if not b_f_stack:
                return i
            b_f_stack.pop()
            if b_f_stack:
                cur_f_char = b_f_stack[len(b_f_stack)-1]
            else:
                cur_f_char = "{"
        elif not _prechar_is_rdline(pre_char) and char in _find_assist_match_pairs:
            if not b_f_stack:
                b_f_stack.append(char)
                cur_f_char = char
            else:
                if _find_assist_enter_stack_pairs[b_f_stack[len(b_f_stack) - 1]][char]:
                    b_f_stack.append(char)
                    cur_f_char = char
        elif not b_f_stack:
            pass
        pre_char = char
    # -1 means no match }
    return -1

##################################################################
def fix_double_quote(file_path,encoding,replace_char=''):
    if not encoding:
        encoding = 'utf-8'
    content = readFile(file_path)
    lines = content.split("\n")
    processed_lines = [process_line(line,replace_char) for line in lines]

    # 写入处理后的内容到输出文件
    writeFile(file_path,'\n'.join(processed_lines))

def process_line(line,replace_char):
    # 统计引号的数量
    quote_count = line.count('"')
    
    # 如果引号数量超过2个
    if quote_count > 2:
        # 找到第一个引号的位置
        first_quote_idx = line.find('"')
        # 找到最后一个引号的位置
        last_quote_idx = line.rfind('"')
        
        # 构造新的行，首尾保留引号，其它引号替换为空字符
        new_line = line[:first_quote_idx+1] + line[first_quote_idx+1:last_quote_idx].replace('"', replace_char) + line[last_quote_idx:]
        return new_line
    else:
        return line
##################################################################

pattern_ch = re.compile(r'[\u4e00-\u9fff]')
def has_chinese(s):
    return bool(pattern_ch.search(s))

##################################################################
# 统计字符串中的格式化数量
pattern_format = re.compile(r'(\%)(.*?)([%-/dfscox])')
def _calc_str_format_map(format_str):
    collect_map = {}
    serial_info = ''
    searchObject = re.findall(pattern_format,format_str)
    total_count = 0
    if searchObject and len(searchObject) > 0:
        for obj in searchObject:
            format_char = obj[0]
            extra_control = obj[1]
            format_type = obj[2]
            if format_type == '%' or format_type == '/' or format_type == '-':
                continue
            params = None
            if format_type in collect_map:
                params = collect_map[format_type]
            else:
                params = {g_count_key:0,g_list_key:[]}
                collect_map[format_type] = params
            
            params[g_count_key] = params[g_count_key] + 1
            params[g_list_key].append(f"{format_char}{extra_control}{format_type}")
            total_count = total_count + 1
            serial_info = serial_info + format_type

    if total_count > 0:
        # 没有就不拼装key了
        collect_map[g_count_key] = total_count
        collect_map[g_serial_key] = serial_info
        collect_map[g_all_same_key] = len(set(serial_info)) == 1



    return collect_map
##################################################################

def readTranslatedExcel(file_ab_path,exported_map,exported_list):
    # file_path = os.path.abspath(g_mother_en_path + "/../" + file_name)
    ret_map = exported_map
    assist_tb_id_map = {}
    # global g_origin_translate_result_map
    if os.path.exists(file_ab_path):
        # 创建一个 Excel 文件
        workbook = openpyxl.load_workbook(file_ab_path)
        worksheet = workbook.active
        data_start_row = -1
        can_collect_datas = False
        row_index = -1
        row_head = None
        for row in worksheet.iter_rows(values_only=True):
            row_index = row_index + 1
            if not can_collect_datas:
                if row[0] == "table":
                    row_head = row
                    can_collect_datas = True
                    # row 的 lang key 列必须有值：
                    if data_start_row == -1:
                        data_start_row = row_index + 1
            else:
                table_name = row[0]
                if table_name == None:
                    continue
                table_name = table_name.lower()
                id = str(row[1])
                key = row[2]
                en = row[3]
                map = None
                keys = None
                struct = None

                table_name_id_union_key = f"{table_name}^{id}^{key}"

                if table_name in ret_map:
                    map = ret_map[table_name]
                else:
                    map = {}
                    ret_map[table_name] = map
    
                if id != None and id != str(None):
                    if id in map:
                        keys = map[id]
                    else:
                        keys = {}
                        map[id] = keys
            
                if key != None:
                    if key in keys:
                        struct = keys[key]
                    else:
                        struct = {}
                        keys[key] = struct
                
                if struct and g_en_key in struct:
                    pass
                else:
                    if en:
                        struct["en"] = en

                cache_tb_map = None
                if table_name_id_union_key in assist_tb_id_map:
                    cache_tb_map = assist_tb_id_map[table_name_id_union_key]
                else:
                    cache_tb_map = {g_table_name_key:table_name,g_id_key:id,g_sort_key:key}
                    assist_tb_id_map[table_name_id_union_key] = assist_tb_id_map
                    if exported_list:
                        exported_list.append(cache_tb_map)
    
                # 每个文件里可能有多个语言列
                lang_l = len(row)
                for i in range(4,lang_l):
                    lang = row_head[i]
                    if not lang:
                        continue
                    lang_str = row[i]
                    if lang_str:
                        lang_str = lang_str.replace('"',g_double_quote_replace_char)
                        struct[lang] = lang_str

                        # 原始翻译结果key
                        # if lang != g_en_key:
                        #     g_origin_translate_result_map[struct[g_en_key] + "_" + lang] = lang_str

                        unit_key = f"{table_name}^{id}^{key}"
                        if unit_key in g_dup_map_from_origin:
                            g_dup_map_from_origin[unit_key]["main"] = unit_key
                        

                    if lang not in g_read_langs:
                        g_read_langs[lang] = lang

                    langs_map = None
                    if key in cache_tb_map:
                        langs_map = cache_tb_map[key]
                    else:
                        langs_map = {}
                        cache_tb_map[key] = langs_map
                    langs_map[lang] = lang_str

    # 返回树状解析结构
    return ret_map

patternOneWordSub = re.compile(r'\s*(\w+)\s*', re.M|re.S)
patternTranslateSub = re.compile(r'\s*(\w+)\s*=\s*"(.*?)"\s*;', re.M|re.S)
# 从分段字符串中检索
# 纯收集，注意这里没有替换
def addToTranslateCheck(tableName,content,decl_vars,sub_struct_info,current_lang_key,collect_map):
    to_search_content = content
    
    searchOneWorldObject = re.search(patternOneWordSub,to_search_content)
    id = ""
    trans_struct = None
    new_create = False
    is_mother_table = current_lang_key == g_en_key

    if not sub_struct_info:
        sub_struct_info = decl_vars
    else:
        # 有子类型，先定义结构体
        new_create = True
        trans_struct = collect_map

    next_step_count = 0

    # 数组字段
    var_collect_list = None

    while searchOneWorldObject:
        word = searchOneWorldObject.group(1)
        if word in sub_struct_info:
            var_info = sub_struct_info[word]

            # 准备数组
            if var_info[g_analyse_var_if_array_key]:
                # 多语言收集时忽略数组
                if g_no_lang_mode:
                    if word in trans_struct:
                        var_collect_list = trans_struct[word]
                    else:
                        var_collect_list = []
                        trans_struct[word] = var_collect_list
            else:
                var_collect_list = None

            # 如果是复杂结构
            if var_info[g_analyse_var_if_custom_key]:
                # 递归分析结构内容

                var_collect_map = {}
                var_type_str = var_info[g_analyse_var_type_key]
                target_struct_info = decl_vars[g_decl_struct_key][var_type_str]
                str_content,step_count = doSubStructString(to_search_content)
                doCollectToTrnaslateList(str_content,decl_vars,target_struct_info,current_lang_key,var_collect_list,None,var_collect_map)
                next_step_count = step_count

                if word not in trans_struct:
                    trans_struct[word] = var_collect_map
                pass
            else:
                # 简单结构使用 a=b匹配
                searchObject = re.search(patternTranslateSub,to_search_content)
                next_step_count = searchObject.end(0)
                
                vVarName = searchObject.group(1)
                vVarValue = searchObject.group(2)
                if id == "" and vVarName == g_id_key:
                    id = vVarValue
                    global_key = f"{tableName}^{id}"
                    if global_key in collect_map:
                        trans_struct = collect_map[global_key]
                    else:
                        trans_struct = {}
                        trans_struct[g_table_name_key] = tableName
                        trans_struct[g_id_key] = id
                        collect_map[global_key] = trans_struct

                        if g_no_lang_mode:
                            # 普通表这里加个缓存优化
                            all_table_content = None
                            if tableName in collect_map:
                                all_table_content = collect_map[tableName]
                            else:
                                all_table_content = {}
                                collect_map[tableName] = all_table_content
                            all_table_content[id] = trans_struct

                        new_create = True

                if vVarName in sub_struct_info:
                    # 此时 word 就是 vVarName
                    collect_the_var = True
                    as_static_var = False
                    # 多语言时要处理为静态数据
                    if not g_no_lang_mode and var_info[g_analyse_var_ignore_key]:
                        as_static_var = True
                    if not g_no_lang_mode:
                        if vVarValue.isdigit() or not var_info[g_analyse_var_is_pure_str_key]:
                            as_static_var = True
                    # 收集普通表时不能忽略，因为要回写
                    if not g_no_lang_mode:
                        # 多语言表模式下，要忽略几个语言字段
                        if tableName in g_const_strs:
                            const_tb = g_const_strs[tableName]
                            # id 在前面，一定已经处理过
                            if id in const_tb[g_id_key] and vVarName in const_tb["const_key"]:
                                # collect_the_var = False
                                as_static_var = True

                    if collect_the_var:
                        global_key = f"{tableName}^{id}^{vVarName}"
                        
                        # 是否要跳过
                        if g_no_lang_mode:
                            if var_info[g_analyse_var_if_array_key]:
                                var_collect_list.append(vVarValue)
                            else:
                                # 散字段似乎没有什么必要
                                # collect_map[global_key] = vVarValue
                                trans_struct[vVarName] = vVarValue
                                pass
                            
                        else:
                            if as_static_var:
                                static_var_map = None
                                if g_static_var_key in trans_struct:
                                    static_var_map = trans_struct[g_static_var_key]
                                else:
                                    static_var_map = {}
                                    trans_struct[g_static_var_key] = static_var_map
                                if vVarName in static_var_map:
                                    if vVarValue != static_var_map[vVarName]:
                                        if vVarName != g_name_key:
                                            logger.error(f"tb:{tableName} lang:{current_lang_key} id:{id} static_values not match!!!! part name:{vVarName} exist:{static_var_map[vVarName]} new:{vVarValue}.")
                                            all_need_collector_log.append(f"tb:{tableName} lang:{current_lang_key} id:{id} static_values not match!!!! part name:{vVarName} exist:{static_var_map[vVarName]} new:{vVarValue}.")
                                else:
                                    static_var_map[vVarName] = vVarValue
                            else:
                                lang_struct = None
                                if global_key in collect_map:
                                    lang_struct = collect_map[global_key]
                                else:
                                    lang_struct = {}
                                    collect_map[global_key] = lang_struct
                                # if is_mother_table:
                                #     if has_chinese(vVarValue):
                                #         lang_struct[g_chs_key] = vVarValue
                                #     else:
                                #         lang_struct[g_en_key] = vVarValue
                                # else:
                                #     lang_struct[current_lang_key] = vVarValue
                                if len(current_lang_key) > 0:
                                    lang_struct[current_lang_key] = vVarValue
                                    trans_struct[vVarName] = lang_struct
                                # else:
                                #     trans_struct[vVarName] = vVarValue


        else:
            # if word != tableName:
            logger.error(f"addToTranslateCheck error:word {word} not in sub_struct_info from map {tableName}, current lang:{current_lang_key}, whole info:{to_search_content}.")
            all_need_collector_log.append(f"addToTranslateCheck error:word {word} not in sub_struct_info from map {tableName}, current lang:{current_lang_key}, whole info:{to_search_content}.")
            break
        to_search_content = to_search_content[next_step_count:]
        searchOneWorldObject = re.search(patternOneWordSub,to_search_content)
    return trans_struct,new_create

# 从一个字符串中取前面的结构体字符串，返回字符串和此方式步进的字符数
def doSubStructString(content):
    searchObject = re.search(patternToTranslateList,content)
    step_count = 0
    while searchObject:
        vTableName = searchObject.group(1)
        cut_content = content[searchObject.end(0):]
        r_index = find_right_f_edge(cut_content)
        if r_index == -1:
            logger.error(f"doSubStructString error found no edge for {vTableName} near: {cut_content[:log_near_char_count]}")
            all_need_collector_log.append(f"doSubStructString error found no edge for {vTableName} near: {cut_content[:log_near_char_count]}")
        else:
            # vars_string = cut_content[:r_index]
            step_count = searchObject.end(0) + r_index + 1
            full_ret = content[:step_count]
            return full_ret,step_count
    return None,step_count

patternToTranslateList = re.compile(r'^\s*(\w+)\s*\{', re.M|re.I|re.S)
# 如果是 sub_struct_info 为 None ，则是根结点
def doCollectToTrnaslateList(content,decl_vars,sub_struct_info,current_lang_key,collect_list,collect_dup_map,collect_map):
    if content and len(content) > 0:
        to_search_content = content
        # printed = False
        searchObject = re.search(patternToTranslateList,to_search_content)
        while searchObject:
            vTableName = searchObject.group(1)
            # if not printed:
            #     printed = True
            #     logger.debug(f"table name:{vTableName}")
            cut_content = to_search_content[searchObject.end(0):]
            r_index = find_right_f_edge(cut_content)
            if r_index == -1:
                logger.error(f"doCollectToTrnaslateList error found no edge for {vTableName} when collect lang {current_lang_key} at here:{cut_content[0:20]}")
                all_need_collector_log.append(f"doCollectToTrnaslateList error found no edge for {vTableName} when collect lang {current_lang_key} at here:{cut_content[0:20]}")
            else:
                vars_string = cut_content[:r_index]
                trans_struct, new_create = addToTranslateCheck(vTableName,vars_string,decl_vars,sub_struct_info,current_lang_key,collect_map)
                # 子类型不添加到全数据列表数组
                # if not g_no_lang_mode:
                # if not sub_struct_info:
                if collect_list != None and new_create:
                    collect_list.append(trans_struct)

                if not sub_struct_info and collect_dup_map and trans_struct:
                    id_map = collect_dup_map[g_id_key]
                    tb_map = collect_dup_map[g_table_name_key]
                    id = trans_struct[g_id_key]
                    if id in id_map:
                        dup_arr = collect_dup_map[g_dup_key]
                        dup_info_str = f"dup item:{vTableName} {id}, already exist at {id_map[id][g_table_name_key]}"
                        dup_arr.append(dup_info_str)
                        # 重复项出现
                        dup_tb_arr = None
                        if vTableName in tb_map:
                            dup_tb_arr = tb_map[vTableName]
                        else:
                            dup_tb_arr = []
                            tb_map[vTableName] = dup_tb_arr
                        dup_tb_arr.append(trans_struct)
                    else:
                        id_map[id] = {g_table_name_key:vTableName,g_table_data_name_key:trans_struct}

            
            to_search_content = to_search_content[searchObject.end(0) + r_index:]
            searchObject = re.search(patternToTranslateList,to_search_content)


def doTest():
    # DLC_NODES
    content = readFile("D:\\work\\fishing\\svn\\Release_global\\datapool\\ElementData\\BaseData\\ACHIEVEMENT_WANTED.decl.h","latin1")
    decl_vars_map = doCollectDeclMap("ACHIEVEMENT_WANTED",content)
    datas_content = readFile("D:\\work\\fishing\\svn\\Release_global\\datapool\\ElementData\\BaseData\\ACHIEVEMENT_WANTED.data.txt")
    collect_list = []
    collect_map = {}
    doCollectToTrnaslateList(datas_content,decl_vars_map,None,"",collect_list,None,collect_map)
    pass

# 分析数据结构，可能嵌套
# 根据格式，解析出每个变量的信息，包含类型
# 兼容旧的使用方式，根字段为外层结构体的原始字段
# 与table_name 不同的结构体为 依赖结构体
def doAnalyseDataStruct(table_name,content,export_map):
    searchObject = re.findall(patternAnyVarsList,content)
    if searchObject and len(searchObject) > 0:
        array = []
        export_map[g_decl_var_order_key] = array
        index = 0
        for obj in searchObject:
            var_type = obj[0]
            var_name = obj[1]
            var_array = obj[2]
            var_issue = obj[3]
            # if var_name == "name":
            #     # 名称字段在任意模式下都被忽略
            #     pass
            # else:
            var_info = {}
            var_info[g_analyse_var_type_key] = var_type
            var_info[g_analyse_var_name_key] = var_name
            var_info[g_analyse_var_if_array_key] = len(var_array) >= 2
            var_info[g_analyse_var_issue_key] = var_issue
            var_info[g_analyse_var_is_pure_str_key] = var_type == "string" and not var_info[g_analyse_var_if_array_key]
            var_info[g_analyse_var_if_custom_key] = var_type != "int" and var_type != "string" and var_type != "float"
            var_info[g_analyse_var_ignore_key] = var_name == "name" # 因为涉及回写功能，所有字段全部收集
            var_info[g_analyse_var_index_key] = index

            index = index + 1

            export_map[var_name] = var_info
            array.append(var_info)
            

patternCollectDeclList = re.compile(r'^\s*struct\s+(\w+)\s*(\/\/.*?\n)?\{', re.M|re.I|re.S)
patternStrVarsList = re.compile(r'^\s*string\s+(\w+)\s*;', re.M|re.I|re.S)
patternAnyVarsList = re.compile(r'^\s*(\w+)\s+(\w+)\s*(\[\s*\])?\s*;?(.*?)\n', re.M|re.I|re.S)
def doCollectDeclMap(table_name_up,content):
    all_struct_map = {}
    decl_vars = {g_decl_struct_key:all_struct_map}
    if content and len(content) > 0:
        to_search_content = content
        printed = False
        searchObject = re.search(patternCollectDeclList,to_search_content)
        while searchObject:
            vTableName = searchObject.group(1)
            v_stuct_map = {}
            all_struct_map[vTableName] = v_stuct_map
            if not printed:
                printed = True
                logger.debug(f"decl table name:{vTableName}")
            cut_content = to_search_content[searchObject.end(0):]
            r_index = find_right_f_edge(cut_content)
            if r_index == -1:
                logger.error(f"doCollectDeclList error found no edge for {vTableName} near: {cut_content[:log_near_char_count]}")
                all_need_collector_log.append(f"doCollectDeclList error found no edge for {vTableName} near: {cut_content[:log_near_char_count]}")
            else:
                vars_string = cut_content[:r_index]
                doAnalyseDataStruct(vTableName,vars_string,v_stuct_map)

            to_search_content = to_search_content[searchObject.end(0) + r_index:]
            searchObject = re.search(patternCollectDeclList,to_search_content)
    if table_name_up in all_struct_map:
        main_struct = all_struct_map[table_name_up]
        for var in main_struct:
            decl_vars[var] = main_struct[var]
    else:
        logger.error(f" can't find main struct as {table_name_up}")
        all_need_collector_log.append(f" can't find main struct as {table_name_up}")
    return decl_vars

# file_ab_path txt文件路径
# current_lang_key 所属语言，普通表填 en 就行
# collect_list 以数组形式收集，可不填
# collect_dup_map 是否收集重复内容，可不填
# collect_map 以map形式收集，必填
def doWorkOnDataTxtFile(file_ab_path,current_lang_key,collect_list,collect_dup_map,collect_map):
    table_name_up = os.path.basename(file_ab_path).replace(".data.txt","")
    decl_path = file_ab_path.replace(".data.txt",".decl.h")
    decl_content = readFile(decl_path,"latin1")
    decl_vars = doCollectDeclMap(table_name_up,decl_content)
    if decl_vars and len(decl_vars) > 0:
        content = readFile(file_ab_path)
        if content and len(content) > 0:
            doCollectToTrnaslateList(content,decl_vars,None,current_lang_key,collect_list,collect_dup_map,collect_map)
    else:
        logger.error(f"doWorkOnDataTxtFile decl {decl_path} no vars to translate, need check.")
        all_need_collector_log.append(f"doWorkOnDataTxtFile decl {decl_path} no vars to translate, need check.")

# table_classify 参数决定收集哪些表分类文件
def doCollectFromMotherDatas(table_classify: ClassifyFile,mother_en_path,current_lang_key,collect_list,collect_dup_map,collect_map,on_table_names_map=None):
    if mother_en_path and len(mother_en_path) > 0:
        for root, _, files in os.walk(mother_en_path):
            for file in files:
                file_ab_path = os.path.join(root, file)
                table_name_up = file.replace(".data.txt","")
				# test one file
                # if file != "FISHERIES_ACTIVITY_LANGUAGE.data.txt":
                #     continue
                # test end
                # add test mode for quick debug
                if g_test_tables and file not in g_test_tables:
                    continue
                if on_table_names_map:
                    # 需要在里面，不在就下一个
                    if not table_name_up.lower() in on_table_names_map:
                        continue
                if table_classify == ClassifyFile.NormalTables:
                    if not file.endswith("_LANGUAGE.data.txt") and file.endswith(".data.txt"):
                        doWorkOnDataTxtFile(file_ab_path,current_lang_key,collect_list,collect_dup_map,collect_map)
                elif table_classify == ClassifyFile.LangTables:
                    if file.endswith("_LANGUAGE.data.txt"):
                        doWorkOnDataTxtFile(file_ab_path,current_lang_key,collect_list,collect_dup_map,collect_map)
                elif table_classify == ClassifyFile.AllTables:
                    if file.endswith(".data.txt"):
                        doWorkOnDataTxtFile(file_ab_path,current_lang_key,collect_list,collect_dup_map,collect_map)

# collect_dup_map 是收集到的重复id数据
# on_table_names_map 是在指定的表上收集
def doWorkCollectOnFolderAndExport(table_classify,mother_en_path,other_lang_path,collect_list,collect_dup_map,collect_map,withTranslated=False,on_table_names_map=None):
    '''从母表中收集多语言，并导出为excel'''
    doCollectFromMotherDatas(table_classify,mother_en_path,g_en_key,collect_list,collect_dup_map,collect_map,on_table_names_map)
    if withTranslated:
        files_and_folders = os.listdir(other_lang_path)
        for item in files_and_folders:
            dir_ab_path = os.path.join(other_lang_path, item)
            if os.path.isdir(dir_ab_path):
                lang = item.lower()
                if lang not in g_read_langs:
                    g_read_langs[lang] = lang
                for root, _, files in os.walk(dir_ab_path):
                    for file in files:
                        file_ab_path = os.path.join(root, file)
                        if file.endswith("_LANGUAGE.data.txt"):
                            # test one file
                            # if file != "FISHERIES_ACTIVITY_LANGUAGE.data.txt":
                            #     continue
                            # test end
                            # add test mode for quick debug
                            table_name_up = file.replace(".data.txt","")

                            if g_test_tables and file not in g_test_tables:
                                continue

                            if on_table_names_map:
                                # 需要在里面，不在就下一个
                                if not table_name_up.lower() in on_table_names_map:
                                    continue

                            doWorkOnDataTxtFile(file_ab_path,lang,collect_list,collect_dup_map,collect_map)

def doCollectAllLanguagesFromElementDataPath(path):
    mother_en_path = os.path.abspath(path + os.sep + "BaseData")
    other_lang_path = os.path.abspath(path + os.sep +  "/../Localization/ElementData/Loc")

    collect_list = []
    collect_map = {}
    # todo 先不收集多语言
    withTranslated = True
    if g_no_lang_mode:
        # 读取普通表时，不需要读取翻译表
        withTranslated = False

    table_classify = ClassifyFile.NoTables
    if g_no_lang_mode:
        table_classify = ClassifyFile.NormalTables
    else:
        table_classify = ClassifyFile.LangTables
    doWorkCollectOnFolderAndExport(table_classify,mother_en_path,other_lang_path,collect_list,None,collect_map,withTranslated)
    return collect_map,collect_list

def doWorkReadTranslatedExcels(container_path):
    files_and_folders = os.listdir(container_path)
    exported_map = {}
    exported_list = []
    for item in files_and_folders:
        full_path = os.path.join(container_path, item)
        if os.path.isfile(full_path) and (item.endswith(".xlsx") and (item.startswith("translated") or item.startswith("to_translate"))):
            readTranslatedExcel(full_path,exported_map,exported_list)
    return exported_map,exported_list

def convert_element_to_string(element):
    
    title = f"{element[g_table_name_key]}^{element[g_id_key]}"
    str_list = []
    for k in element:
        if k != g_id_key and k != g_table_name_key and k != g_sort_key and k != g_static_var_key:
            ele = element[k]
            if g_no_lang_mode:
                str_list.append([k,str(ele)])
            else:
                en_str = ""
                if g_en_key in ele:
                    en_str = ele[g_en_key]
                elif g_chs_key in ele:
                    en_str = ele[g_chs_key]
                if en_str and len(en_str) > 0:
                    # str_list.append(f"{title}^{k}^{en_str}")
                    str_list.append([k,en_str])

    def custom_compare(data1, data2):
        # 表名字母排序
        if data1[0] < data2[0]:
            return -1
        elif data1[0] > data2[0]:
            return 1
        else:
            return 0

    key_function = cmp_to_key(custom_compare)
    str_list = sorted(str_list, key=key_function)
    modified_arr = str_list[:]
    for index, old_e in enumerate(str_list):
        modified_arr[index] = f"{title}^{old_e[0]}^{old_e[1]}"

    return "\n".join(modified_arr)

def saveToTxt(file_path,custom_list):
    # 转换为字符串
    str_list = [convert_element_to_string(element) for element in custom_list]

    result_string = "\n".join(str_list)
    writeFile(file_path,result_string)

def decompressMaps(version_map,version_list,dup_map):

    deal_assist = {}
    list_assist = {}
    for k in dup_map:
        dup_map_data = dup_map[k]
        key_list = dup_map_data["list"]
        if len(key_list) > 1:
            main_key = dup_map_data["main"]
            if main_key in deal_assist:
                pass
            else:
                deal_assist[main_key] = key_list
                main_key_arr = main_key.split('^')
                table_name = main_key_arr[0]
                id = main_key_arr[1]
                part = main_key_arr[2]

                main_lang_info = version_map[table_name][id][part]

                for dup_key in key_list:
                    if dup_key != main_key:
                        # 需要追加的key信息
                        dup_key_arr = dup_key.split('^')
                        dup_table_name = dup_key_arr[0]
                        dup_id = dup_key_arr[1]
                        dup_part = dup_key_arr[2]

                        if dup_table_name in version_map:
                            tb_ = version_map[dup_table_name]
                            if dup_id in tb_:
                                if dup_part in tb_[dup_id]:
                                    pass
                                else:
                                    tb_[dup_id][dup_part] = main_lang_info
                            else:
                                tb_[dup_id] = {dup_part:main_lang_info}
                            list_assist_key = f"{dup_table_name}^{dup_id}"
                            list_assist_struct = None
                            if list_assist_key in list_assist:
                                list_assist_struct = list_assist[list_assist_key]
                            else:
                                list_assist_struct = {g_table_name_key:dup_table_name,g_id_key:dup_id,g_sort_key:dup_part}
                                version_list.append(list_assist_struct)
                            list_assist_struct[dup_part] = main_lang_info
                        else:
                            logger.error(f"version_map not contains table:{dup_table_name}")
                            all_need_collector_log.append(f"version_map not contains table:{dup_table_name}")

def sortVersionList(version_list):
    # 自定义比较函数
    def custom_compare(data1, data2):
        # 表名字母排序
        if data1[g_table_name_key] < data2[g_table_name_key]:
            return -1
        elif data1[g_table_name_key] > data2[g_table_name_key]:
            return 1
        else:
            # 然后按id排序
            id1 = int(data1[g_id_key])
            id2 = int(data2[g_id_key])
            if id1 != id2:
                return id1 - id2
            else:
                if data1[g_sort_key] < data2[g_sort_key]:
                    return -1
                elif data1[g_sort_key] > data2[g_sort_key]:
                    return 1
                else:
                    return 0

    # 使用 cmp_to_key 将比较函数转换为 key 函数
    key_function = cmp_to_key(custom_compare)

    # 使用 sorted() 函数进行自定义排序
    version_list = sorted(version_list, key=key_function)
    return version_list

def regerenateVersionList(version_map):
    ret_list = []
    for table_name in version_map:
        table_info = version_map[table_name]
        for id in table_info:
            id_data = table_info[id]
            for part in id_data:
                ret_list.append({g_table_name_key:table_name,g_id_key:id,g_sort_key:part,part:id_data[part]})

    return ret_list

def are_dicts_equal(dict1, dict2):
    if len(dict1) != len(dict2):
        return False
    for key in dict1:
        if key not in dict2:
            return False
        if dict1[key] != dict2[key]:
            return False
    return True

# 判定两个id的数据是否完全相同
def ifAllDatasSame(version_map1,version_map2,table_name,id):
    id_data1 = None
    id_data2 = None
    cancat_key = f"{table_name}^{id}"
    if table_name in version_map1:
        table_info = version_map1[table_name]
        if id in table_info:
            id_data1 = table_info[id]
    if not id_data1:
        if cancat_key in version_map1:
            id_data1 = version_map1[cancat_key]

    if table_name in version_map2:
        table_info = version_map2[table_name]
        if id in table_info:
            id_data2 = table_info[id]

    if not id_data2:
        if cancat_key in version_map2:
            id_data2 = version_map2[cancat_key]

    if not id_data1 and not id_data2:
        return True
    if (id_data1 and not id_data2) or (not id_data1 and id_data2):
        return False
    return are_dicts_equal(id_data1,id_data2)

# 参数都是map
def getDataFromVersionMap(version_map,table_name,id,part,lang):
    if table_name in version_map:
        table_info = version_map[table_name]
        if id in table_info:
            id_data = table_info[id]
            if part in id_data:
                part_info = id_data[part]
                if not lang:
                    return part_info
                if lang in part_info:
                    lang_str = part_info[lang]
                    return lang_str

    cancat_key = f"{table_name}^{id}^{part}"
    if cancat_key in version_map:
        part_info = version_map[cancat_key]
        if not lang:
            return part_info

        if lang in part_info:
            lang_str = part_info[lang]
            return lang_str
        
    cancat_key = f"{table_name}^{id}"
    if cancat_key in version_map:
        id_tb_info = version_map[cancat_key]
        if part in id_tb_info:
            part_info = id_tb_info[part]
            if not lang:
                return part_info
            if lang in part_info:
                lang_str = part_info[lang]
                return lang_str
    return None
    
def doCompareThreeVersionDetails(old_510,china_712,new_810,write_path,old_version_path,china_712_path,new_version_path):
    new_to_translate = []
    export_both_modified_list = []
    need_to_write_back_map = {}
    if False:
        # 从配置表读取
        # 遍历 810版本最新内容
        for table_name in new_810:
            table_info = new_810[table_name]
            for id in table_info:
                id_data = table_info[id]
                for part in id_data:
                    # 810版本的最新内容
                    lang_810_str = id_data[part][g_en_key]
                    # 510版本的内容
                    lang_510_str = getDataFromVersionMap(old_510,table_name,id,part,g_en_key)
                    # 712版本的内容
                    lang_712_str = getDataFromVersionMap(china_712,table_name,id,part,g_en_key)
                    if lang_810_str != lang_510_str:
                        # 810和510有差异
                        if lang_510_str == lang_712_str:
                            # 但 510和712无差异
                            # 要覆盖的情况
                            pass
                        else:
                            # 510和712有差异
                            # 导出独立excel，列出原内容和两边的修改内容
                            export_both_modified_list.append({g_table_name_key:table_name,g_id_key:id,g_sort_key:part,part:{g_old_510_key:lang_510_str,g_china_712_key:lang_712_str,g_global_810_key:lang_810_str}})
                            pass
    else:
        # 遍历 810版本最新内容
        for key_pair in new_810:
            table_info = new_810[key_pair]
            if g_table_name_key in table_info and g_id_key in table_info and len(table_info) > 2:
                for part in table_info:
                    if part != g_table_name_key and part != g_id_key and part != g_static_var_key:

                        if g_no_lang_mode:
                            # 普通表时
                            table_name = table_info[g_table_name_key]
                            id = table_info[g_id_key]
                            part_info = table_info[part]
                            # 810版本的最新内容
                            lang_810_str = str(part_info)
                            # 712 版本的内容
                            data_712 = getDataFromVersionMap(china_712,table_name,id,part,None)
                            lang_712_str = str(data_712)
                            # 510版本的内容
                            data_510 = getDataFromVersionMap(old_510,table_name,id,part,None)
                            lang_510_str = str(data_510)
                            if lang_810_str != lang_510_str:
                                # 810和510有差异
                                if lang_510_str == lang_712_str:
                                    # 但 510和712无差异
                                    # 要覆盖的情况
                                    # 需要整个表全一致才会回写
                                    # if ifAllDatasSame(china_712,old_510,table_name,id):
                                    write_back_struct = None
                                    if key_pair in need_to_write_back_map:
                                        write_back_struct = need_to_write_back_map[key_pair]
                                    else:
                                        write_back_struct = {g_table_data_name_key:table_info}
                                        need_to_write_back_map[key_pair] = write_back_struct
                                    write_back_struct[part] = part_info
                                    pass
                                # 需要810和712也不同，再导出
                                elif lang_810_str != lang_712_str:
                                    # 510和712有差异
                                    # 导出独立excel，列出原内容和两边的修改内容
                                    export_both_modified_list.append({g_table_name_key:table_name,g_id_key:id,g_sort_key:part,part:{g_old_510_key:lang_510_str,g_china_712_key:lang_712_str,g_global_810_key:lang_810_str}})
                                    pass
                        else:
                            # 多语言时
                            table_name = table_info[g_table_name_key]
                            id = table_info[g_id_key]
                            part_info = table_info[part]
                            # 810版本的最新内容
                            lang_810_str = ""
                            if g_en_key in part_info:
                                lang_810_str = part_info[g_en_key]
                            # 712 版本的内容
                            lang_712_str = getDataFromVersionMap(china_712,table_name,id,part,g_en_key)
                            # 510版本的内容
                            lang_510_str = getDataFromVersionMap(old_510,table_name,id,part,g_en_key)
                            if lang_810_str != lang_510_str:
                                # 810和510有差异
                                if lang_510_str == lang_712_str:
                                    # 但 510和712无差异
                                    # 要覆盖的情况
                                    # 需要整个表全一致才会回写
                                    # if ifAllDatasSame(china_712,old_510,table_name,id):
                                    # need_to_write_back_map[key_pair] = table_info
                                    write_back_struct = None
                                    if key_pair in need_to_write_back_map:
                                        write_back_struct = need_to_write_back_map[key_pair]
                                    else:
                                        write_back_struct = {g_table_name_key:table_name,g_id_key:id}
                                        if g_static_var_key in table_info:
                                            write_back_struct[g_static_var_key] = table_info[g_static_var_key]
                                        need_to_write_back_map[key_pair] = write_back_struct
                                    write_back_struct[part] = part_info
                                    # else:
                                    #     pass
                                    pass
                                # 需要810和712也不同，再导出
                                elif lang_810_str != lang_712_str:
                                    # 510和712有差异
                                    # 导出独立excel，列出原内容和两边的修改内容
                                    export_both_modified_list.append({g_table_name_key:table_name,g_id_key:id,g_sort_key:part,part:{g_old_510_key:lang_510_str,g_china_712_key:lang_712_str,g_global_810_key:lang_810_str}})
                                    pass

    # 遍历 712 版本最新内容
    for key_pair in china_712:
        table_info = china_712[key_pair]
        if g_table_name_key in table_info and g_id_key in table_info and len(table_info) > 2:
            for part in table_info:
                if part != g_table_name_key and part != g_id_key and part != g_static_var_key:

                    if g_no_lang_mode:
                        # 普通表时
                        table_name = table_info[g_table_name_key]
                        id = table_info[g_id_key]
                        part_info = table_info[part]
                        # 712 版本的内容
                        lang_712_str = str(part_info)
                        # 810版本的最新内容
                        lang_810_ori = getDataFromVersionMap(new_810,table_name,id,part,None)
                        lang_810_str = str(lang_810_ori)
                        # 510版本的内容
                        lang_510_str = str(getDataFromVersionMap(old_510,table_name,id,part,None))
                        if lang_712_str != lang_510_str:
                            # 712和510有差异
                            if lang_810_str == lang_510_str:
                                # 但 810和712无差异
                                # 都需要进行本地化，导出另一个独立excel
                                # new_to_translate.append({g_table_name_key:table_name,g_id_key:id,g_sort_key:part,part:part_info})
                                # 按理说，算新增的，不用处理
                                pass
                        else:
                            # 712和510相同，但810不同的内容，需要回写。此种情况下理论上 810应该能遍历到，810没有遍历到的情况，是需要删除字段的情况
                            if lang_810_str != lang_510_str:
                                write_back_struct = None
                                if key_pair in need_to_write_back_map:
                                    write_back_struct = need_to_write_back_map[key_pair]
                                else:
                                    write_back_struct = {g_table_data_name_key:table_info}
                                    need_to_write_back_map[key_pair] = write_back_struct
                                write_back_struct[part] = lang_810_ori
                    else:
                        # 多语言时
                        table_name = table_info[g_table_name_key]
                        id = table_info[g_id_key]
                        part_info = table_info[part]
                        # 712 版本的内容
                        lang_712_str = ""
                        if g_en_key in part_info:
                            lang_712_str = part_info[g_en_key]
                        # 810版本的最新内容
                        lang_810_str = getDataFromVersionMap(new_810,table_name,id,part,g_en_key)
                        # 510版本的内容
                        lang_510_str = getDataFromVersionMap(old_510,table_name,id,part,g_en_key)
                        if lang_712_str != lang_510_str:
                            # 712和510有差异
                            if lang_810_str == lang_510_str:
                                # 但 810和712无差异
                                # 都需要进行本地化，导出另一个独立excel
                                new_to_translate.append({g_table_name_key:table_name,g_id_key:id,g_sort_key:part,part:part_info})
                                pass
                        else:
                            if lang_810_str != lang_510_str:
                                    write_back_struct = None
                                    if key_pair in need_to_write_back_map:
                                        write_back_struct = need_to_write_back_map[key_pair]
                                    else:
                                        write_back_struct = {g_table_name_key:table_name,g_id_key:id}
                                        if g_static_var_key in table_info:
                                            write_back_struct[g_static_var_key] = table_info[g_static_var_key]
                                        need_to_write_back_map[key_pair] = write_back_struct
                                    # 把 810 更新的内容写入
                                    table_info_810 = new_810[key_pair]
                                    if table_info_810 and part in table_info_810:
                                        part_info = table_info_810[part]
                                        write_back_struct[part] = part_info

                    

    export_both_modified_list = sortVersionList(export_both_modified_list)
    writeExportBothModifiedListToExcel3(write_path  + os.sep + "manual.xlsx",export_both_modified_list,old_version_path,china_712_path,new_version_path)

    if not g_no_lang_mode:
        new_to_translate = sortVersionList(new_to_translate)
        writeNeedTranslateListToExcel(write_path  + os.sep + "to_translate_list.xlsx",new_to_translate)
    
    if not g_no_lang_mode:
        convertLangWriteBackMap(need_to_write_back_map)
        # 原理不同，语言表是替换翻译内容的方式
        writeBackLangToVersion(g_china_712_path,need_to_write_back_map)
    else:
        # 普通配置表是全数据回写，与项目组的 excel 功能相同
        china_712
        for key_pair in need_to_write_back_map:
            to_write_datas = need_to_write_back_map[key_pair]
            if key_pair in china_712:
                origin_datas = china_712[key_pair]
                for to_write_key in to_write_datas:
                    origin_datas[to_write_key] = to_write_datas[to_write_key]
            else:
                append_info = to_write_datas[g_table_data_name_key]
                id = append_info[g_id_key]
                tb = append_info[g_table_name_key]
                china_712[key_pair] = append_info
                if tb in china_712:
                    china_712[tb][id] = append_info
                    # logger.error(f"key_pair:{key_pair} not in china_712 version! may be a new one, just append it.")
                    # all_need_collector_log.append(f"key_pair:{key_pair} not in china_712 version! may be a new one, just append it.")
                else:
                    logger.error(f"table {tb} not exist at target version path, please manual add.")
                    all_need_collector_log.append(f"table {tb} not exist at target version path, please manual add.")
        convertNormalTableWriteBackMap(need_to_write_back_map,china_712)
        writeBackNormalTableToVersion(g_china_712_path,need_to_write_back_map,china_712)
        pass

def doCompareTwoVersionDetails(old_510,new_810,write_path):
    # new_to_translate = []
    export_both_modified_list = []
    # need_to_write_back_map = {}
    # 遍历 810版本最新内容
    for key_pair in new_810:
        table_info = new_810[key_pair]
        if g_table_name_key in table_info and g_id_key in table_info and len(table_info) > 2:
            for part in table_info:
                if part != g_table_name_key and part != g_id_key and part != g_static_var_key:

                    if g_no_lang_mode:
                        # 普通表时
                        table_name = table_info[g_table_name_key]
                        id = table_info[g_id_key]
                        part_info = table_info[part]
                        # 810版本的最新内容
                        lang_810_str = str(part_info)
                        # 510版本的内容
                        data_510 = getDataFromVersionMap(old_510,table_name,id,part,None)
                        lang_510_str = str(data_510)
                        if lang_810_str != lang_510_str:
                            # 810和510有差异
                            # 导出独立excel，列出原内容和两边的修改内容
                            export_both_modified_list.append({g_table_name_key:table_name,g_id_key:id,g_sort_key:part,part:{g_old_key:lang_510_str,g_new_key:lang_810_str}})
                            pass
                    else:
                        # 多语言时
                        table_name = table_info[g_table_name_key]
                        id = table_info[g_id_key]
                        part_info = table_info[part]
                        # 810版本的最新内容
                        lang_810_str = ""
                        if g_en_key in part_info:
                            lang_810_str = part_info[g_en_key]
                        # 510版本的内容
                        lang_510_str = getDataFromVersionMap(old_510,table_name,id,part,g_en_key)
                        if lang_810_str != lang_510_str:
                            # 810和510有差异
                            # 导出独立excel，列出原内容和两边的修改内容
                            export_both_modified_list.append({g_table_name_key:table_name,g_id_key:id,g_sort_key:part,part:{g_old_key:lang_510_str,g_new_key:lang_810_str}})
                            pass

    export_both_modified_list = sortVersionList(export_both_modified_list)
    writeExportBothModifiedListToExcel2(write_path  + os.sep + "manual2.xlsx",export_both_modified_list,'旧版本值','新版本值')

    # if not g_no_lang_mode:
    #     new_to_translate = sortVersionList(new_to_translate)
    #     writeNeedTranslateListToExcel(write_path  + os.sep + "to_translate_list.xlsx",new_to_translate)


def get_column_serial(column, plus):
    # 将列号转换为数字表示
    column_num = 0
    for i, c in enumerate(column):
        column_num += (ord(c) - ord('A') + 1) * (26 ** (len(column) - i - 1))
    column_num += plus
    # 将数字表示转换回列号
    result_column = ""
    while column_num > 0:
        remainder = (column_num - 1) % 26
        result_column = chr(ord('A') + remainder) + result_column
        column_num = (column_num - 1) // 26

    return result_column

def convertLangWriteBackMap(target_map):
    # 把map转换为更方便使用的结构
    all_tables = set()

    table_name_mapping_map = {}

    for key_pair in target_map:
        table_info = target_map[key_pair]
        table_name = table_info[g_table_name_key]
        id = table_info[g_id_key]

        all_tables.add(table_name)

        whole_table_info = None
        table_contains_lang = None
        if table_name in table_name_mapping_map:
            whole_table_info = table_name_mapping_map[table_name]
            table_contains_lang = whole_table_info[g_contain_langs_key]
        else:
            table_contains_lang = set()
            whole_table_info = {g_contain_langs_key:table_contains_lang}
            table_name_mapping_map[table_name] = whole_table_info

        whole_table_info[id] = table_info
        for part in table_info:
            if part != g_table_name_key and part != g_id_key and part != g_static_var_key:
                part_info = table_info[part]
                for lang in part_info:
                    if lang not in table_contains_lang:
                        table_contains_lang.add(lang)

    for table_name in table_name_mapping_map:
        target_map[table_name] = table_name_mapping_map[table_name]

    target_map[g_contain_tables_key] = all_tables

def convertNormalTableWriteBackMap(target_map,version_map):
    # 把map转换为更方便使用的结构
    all_tables = set()

    table_name_mapping_map = {}

    for key_pair in target_map:
        # 从版本表里收集需要回写的表名
        table_info = version_map[key_pair]
        table_name = table_info[g_table_name_key]
        id = table_info[g_id_key]

        all_tables.add(table_name)

        whole_table_info = None
        if table_name in table_name_mapping_map:
            whole_table_info = table_name_mapping_map[table_name]
        else:
            whole_table_info = {}
            table_name_mapping_map[table_name] = whole_table_info

        whole_table_info[id] = table_info

    # for table_name in table_name_mapping_map:
    #     target_map[table_name] = table_name_mapping_map[table_name]

    target_map[g_contain_tables_key] = all_tables


# direct_write True: 直接写， Flase 替换的方式写入
# use_new_analyse_static_var 使用传入数据的静态数据，默认 True 使用原始表的静态数据替换
def writeBackLangToVersion(version_path,exported_map,direct_write=False,use_new_analyse_static_var=False):
    # pass

    all_tables = exported_map[g_contain_tables_key]
    # 遍历要写入的内容，而不是遍历所有再检查是否有要写入的内容
    for table_name in all_tables:
        whole_table_info = exported_map[table_name]
        # 涉及的所有需要写入的语言
        table_contains_lang = whole_table_info[g_contain_langs_key]
        for lang in table_contains_lang:
            target_folder = None
            if lang == g_en_key:
                # 英文，在 BaseData 里找
                target_folder = version_path + os.sep + "BaseData"
            else:
                target_folder = version_path + os.sep + "../Localization/ElementData/Loc" + os.sep + lang.upper()
            target_folder = os.path.abspath(target_folder)
            # 再找表名
            # 这种方式要求表名必须与文件名对应，不如全遍历的灵活，但项目里excel表全部列出就是这样强制对应的
            table_path = target_folder + os.sep + table_name.upper() + ".data.txt"
            if not os.path.exists(table_path):
                # 文件不存在，试试小写的
                table_path = target_folder + os.sep + table_name + ".data.txt"
                if not os.path.exists(table_path):
                    logger.error(f"target path {table_path} can't find both upper and lower.")
                    all_need_collector_log.append(f"target path {table_path} can't find both upper and lower.")
                    table_path = ""
            if len(table_path) > 0:
                if direct_write:
                    ######################################################
                    collect_list = []
                    collect_map = {}
                    collect_dup_map = None
                    # 先读取，把name取出来
                    doWorkOnDataTxtFile(table_path,lang,collect_list,collect_dup_map,collect_map)

                    # to_replace_datas = exported_map[table_name]
                    # for data_key in to_replace_datas:
                    #     if data_key != g_contain_langs_key:
                    #         data = to_replace_datas[data_key]
                    #         if g_id_key in data and g_table_name_key in data:
                    #             id = data[g_id_key]
                    #             orgin_data_key = f"{table_name}^{id}"
                    #             if orgin_data_key in collect_map:
                    #                 target_data = collect_map[orgin_data_key]
                    #                 for part in target_data:
                    #                     if part != g_id_key and part != g_table_name_key and part != g_static_var_key:
                    #                         part_info = target_data[part]
                    #                         if lang in part_info:
                    #                             part_info[lang] = data[part][lang]
                    #             else:
                    #                 collect_map[orgin_data_key] = data
                    ######################################################

                    sorted_values = []
                    for id in whole_table_info:
                        table_item = whole_table_info[id]
                        if g_id_key in table_item and g_table_name_key in table_item:
                            new_data = {}
                            orgin_data_key = f"{table_name}^{id}"
                            orgin_data = None
                            orgin_static_data = None
                            if orgin_data_key in collect_map:
                                orgin_data = collect_map[orgin_data_key]
                                if g_static_var_key in orgin_data:
                                    orgin_static_data = orgin_data[g_static_var_key]
                                if use_new_analyse_static_var:
                                    if g_static_var_key in table_item:
                                        orgin_static_data = table_item[g_static_var_key]
                            
                            for part in table_item:
                                part_info = table_item[part]
                                if part == g_static_var_key:
                                    static_vars = part_info
                                    for var in static_vars:
                                        if var == g_name_key:
                                            if orgin_data and var in orgin_static_data:
                                                new_data[var] = orgin_static_data[var]
                                            else:
                                                new_data[var] = static_vars[var]
                                        else:
                                            new_data[var] = static_vars[var]
                                elif part != g_id_key and part != g_table_name_key and lang in part_info:
                                    new_data[part] = part_info[lang]
                            sorted_values.append(new_data)
                    doWriteConfigTable(target_folder,table_name,sorted_values)
                else:
                    doWriteBackToDataTxtFile(table_name,table_path,exported_map,lang)

# 必须按表结构写
# level 几级数据，决定缩进格式
def doSerialDataToStr(data,decl_vars,decl_table_name,table_var_name,level):
    table_structs = decl_vars[g_decl_struct_key]
    if not decl_table_name in table_structs:
        logger.error(f"doSerialDataToStr tb:{decl_table_name} decl no {decl_table_name}")
        all_need_collector_log.append(f"doSerialDataToStr tb:{decl_table_name} decl no {decl_table_name}")
        return ""
    table_struct = table_structs[decl_table_name]
    var_order = table_struct[g_decl_var_order_key]

    vars_arr = []
    empty_chars = "\t" * level
    empty_var_chars = "\t" * (level + 1)
    for var_i in var_order:
        var_name = var_i[g_analyse_var_name_key]
        if var_name in data:
            value = data[var_name]
            if value == None:
                continue
            if var_i[g_analyse_var_if_array_key]:
                if var_i[g_analyse_var_if_custom_key]:
                    custom_type_name = var_i[g_analyse_var_type_key]
                    for sub_value in value:
                        sub_str = doSerialDataToStr(sub_value,decl_vars,custom_type_name,var_name,level + 1)
                        vars_arr.append(sub_str)
                        pass
                else:
                    for sub_value in value:
                        vars_arr.append(f"{empty_var_chars}{var_name}=\"{sub_value}\";")
            else:
                if var_i[g_analyse_var_if_custom_key]:
                    custom_type_name = var_i[g_analyse_var_type_key]
                    sub_str = doSerialDataToStr(value,decl_vars,custom_type_name,var_name,level + 1)
                    vars_arr.append(sub_str)
                else:
                    vars_arr.append(f"{empty_var_chars}{var_name}=\"{value}\";")

    vars_str = "\n".join(vars_arr)
    if len(vars_arr) > 0:
        vars_str = vars_str + "\n"
    final_str = f"{empty_chars}{table_var_name}{{\n{vars_str}{empty_chars}}};"
    return final_str

def doTestDumpData():
    target_folder = "D:\\work\\fishing\\svn\\Release_global\\datapool\\ElementData\\BaseData"

    test_path = target_folder + "\\EVENT_SPECIAL_SALE.decl.h"
    content = readFile(test_path,"latin1")
    test_pattern = re.compile(r'^\s*(\w+)\s+(\w+)\s*(\[\s*\])?\s*;?(.*?)\n', re.M|re.I|re.S)
    searchObject = re.findall(test_pattern,content)
    if searchObject and len(searchObject) > 0:
        index = 0
        for obj in searchObject:
            var_type = obj[0]
            var_name = obj[1]
            var_array = obj[2]
            var_issue = obj[3]

        pass

    for root, _, files in os.walk(target_folder):
        for file in files:
            if file.endswith(".data.txt.bin"):
                file_ab_path = os.path.join(root, file)
                decl_path = file_ab_path.replace(".data.txt.bin",".decl.h")
                txt_path = file_ab_path.replace(".data.txt.bin",".data.txt")
                content = readFile(decl_path,"latin1")
                decl_vars_map = doCollectDeclMap("CHAMPIONSHIPS_TYPE",content)

                bin_path = file_ab_path
                with open(bin_path, 'rb') as file:
                    bin_datas = pickle.load(file)
                    
                    all_strs = []
                    for bin_data in bin_datas:
                        table_name = bin_data[g_table_name_key]
                        table_name_upper = bin_data[g_table_name_key].upper()

                        all_str = doSerialDataToStr(bin_data,decl_vars_map,table_name_upper,table_name,0)
                        all_strs.append(all_str)
                    total_str = "\n" + "\n\n".join(all_strs)

                    writeFile(txt_path,total_str)

# 替换文件时，保留前面的空白
patternCatchHeadEmptyCharsSub = re.compile(r'^(\s*)\w', re.M|re.S)
# 替换文件时，保留后面的空白
patternCatchTailEmptyCharsSub = re.compile(r'(\s*)$', re.M|re.S)
def doWriteConfigTable(target_folder,table_name,array_datas):
    table_name_upper = table_name.upper()
    table_path = target_folder + os.sep + table_name_upper + ".data.txt"

    # 如果之前的文件前部有很多空白，保留之
    origin_content = readFile(table_path)
    empty_chars = ""
    if len(origin_content) > 0:
        searchObject = re.search(patternCatchHeadEmptyCharsSub,origin_content)
        if searchObject:
            empty_chars = searchObject.group(1)
    # test
    # with open(table_path + ".bin", 'wb') as file:
    #     pickle.dump(array_datas, file)

    decl_path = table_path.replace(".data.txt",".decl.h")
    decl_content = readFile(decl_path,"latin1")
    # 拿到数据结构
    decl_vars = doCollectDeclMap(table_name_upper,decl_content)
    # 序列化
    all_strs = []
    for data in array_datas:
        all_str = doSerialDataToStr(data,decl_vars,table_name_upper,table_name,0)
        all_strs.append(all_str)
    total_str = "\n\n".join(all_strs)
    # 回写，和原来的格式一模一样
    writeFile(table_path,empty_chars + total_str + "\n")

def writeBackNormalTableToVersion(version_path,exported_map,version_map):
    # pass
    all_tables = exported_map[g_contain_tables_key]
    # 遍历要写入的内容，而不是遍历所有再检查是否有要写入的内容
    for table_name in all_tables:
        if table_name in version_map:
            whole_table_info = version_map[table_name]
            sorted_keys = sorted(whole_table_info.keys(), key=lambda x: int(x))
            sorted_values = [whole_table_info[key] for key in sorted_keys]
            doWriteConfigTable(version_path + os.sep + "BaseData",table_name,sorted_values)
        else:
            logger.warning(f"writeBackNormalTableToVersion warning : no table name {table_name} in target path {version_path}, this table write back failed.")
                
patternTranslateRepSub = re.compile(r'(\s*)(\w+)\s*=\s*"(.*?)"\s*;', re.M|re.S)
# 从分段字符串中检索并替换
def doWriteBackToStructContent(table_name,content,exported_map,lang):
    if table_name not in exported_map:
        # 无需替换的表
        return content
    replace_table_info = exported_map[table_name]
    replace_keys_info = None
    final_content = ""
    to_search_content = content
    searchObject = re.search(patternTranslateRepSub,to_search_content)
    id = ""

    replaced_keys = {}
    vEmpty = ""
    while searchObject:
        vEmpty = searchObject.group(1)
        vVarName = searchObject.group(2)
        vVarValue = searchObject.group(3)
        if vVarName == "id":
            id = vVarValue
            # 不在被替换列表里
            if id not in replace_table_info:
                # 无需替换的id
                return content
            else:
                replace_keys_info = replace_table_info[id]
        # logger.debug(f"vToTranslateStr:{vToTranslateStr}")
        do_replace = False
        if replace_keys_info and vVarName in replace_keys_info:
            struct = replace_keys_info[vVarName]
            if isinstance(struct,dict) and lang in struct:
                replace_str = struct[lang]
                replaced_str = searchObject.group().replace(vVarValue,replace_str)
                final_content = final_content + to_search_content[0:searchObject.start(0)] + replaced_str
                do_replace = True
                replaced_keys[vVarName] = True
    
        if not do_replace:
            final_content = final_content + to_search_content[0:searchObject.start(0)] + searchObject.group()

        to_search_content = to_search_content[searchObject.end(0):]
        searchObject = re.search(patternTranslateRepSub,to_search_content)
    else:
        if replace_keys_info:
            for key in replace_keys_info:
                if key != g_id_key and key != g_table_name_key and key != g_static_var_key and key not in replaced_keys:
                    # 没能替换的 key, 追加
                    struct = replace_keys_info[key]
                    if isinstance(struct,dict) and lang in struct:
                        final_content = final_content + "\n" + vEmpty + f"{key}=\"{struct[lang]}\";"
        final_content = final_content + to_search_content
    return final_content

patternToTranslateList = re.compile(r'^\s*(\w+)\s*\{', re.M|re.I|re.S)
def doApplyNewDataToDataTxtFile(content,exported_map,lang):
    final_content = ""
    if content and len(content) > 0:
        to_search_content = content
        printed = False
        searchObject = re.search(patternToTranslateList,to_search_content)
        while searchObject:
            vTableName = searchObject.group(1)
            if not printed:
                printed = True
                logger.debug(f"data table name:{vTableName}")
            cut_content = to_search_content[searchObject.end(0):]
            r_index = find_right_f_edge(cut_content)
            replace_str = ""
            if r_index == -1:
                replace_str = searchObject.group()
                logger.error(f"doApplyNewDataToDataTxtFile error found no edge for {vTableName}")
                all_need_collector_log.append(f"doApplyNewDataToDataTxtFile error found no edge for {vTableName}")
            else:
                vars_string = cut_content[:r_index]
                replace_str = searchObject.group() + doWriteBackToStructContent(vTableName,vars_string,exported_map,lang)
            
            final_content = final_content + to_search_content[0:searchObject.start(0)] + replace_str
            to_search_content = to_search_content[searchObject.end(0) + r_index:]
            searchObject = re.search(patternToTranslateList,to_search_content)
        else:
            final_content = final_content + to_search_content
    return final_content

# 把翻译结果写入对应的表格
def doWriteBackToDataTxtFile(table_name,file_ab_path,exported_map,lang):
    # 要支持追加字段和追加模板数据
    if table_name in exported_map:
        collect_list = []
        collect_map = {}
        collect_dup_map = None
        doWorkOnDataTxtFile(file_ab_path,lang,collect_list,collect_dup_map,collect_map)

        to_replace_datas = exported_map[table_name]
        for data_key in to_replace_datas:
            if data_key != g_contain_langs_key:
                data = to_replace_datas[data_key]
                if g_id_key in data and g_table_name_key in data:
                    id = data[g_id_key]
                    orgin_data_key = f"{table_name}^{id}"
                    if orgin_data_key in collect_map:
                        target_data = collect_map[orgin_data_key]
                        for part in target_data:
                            if part != g_id_key and part != g_table_name_key and part != g_static_var_key:
                                part_info = target_data[part]
                                if lang in part_info:
                                    if part in data:
                                        exist_part = data[part]
                                        if lang in exist_part:
                                            part_info[lang] = exist_part[lang]
                                        else:
                                            logger.warning(f"toreplace table:{table_name} id:{id} part:{part} has no lang:{lang}[{g_code_chinese[lang]}]")
                        # 新增
                        for part in data:
                            if part != g_id_key and part != g_table_name_key and part != g_static_var_key:
                                part_info = data[part]
                                if lang in part_info:
                                    if part not in target_data:
                                        target_data[part] = {lang:part_info[lang]}
                                    else:
                                        target_data[part][lang] = part_info[lang]
                    else:
                        collect_map[orgin_data_key] = data

        normal_map = {}
        # 修正字段为普通表
        for data_key in collect_map:
            data = collect_map[data_key]
            if g_id_key in data and g_table_name_key in data:
                id = data[g_id_key]
                new_data = {g_id_key:id,g_table_name_key:data[g_table_name_key]}
                normal_map[id] = new_data
                for part in data:
                    if part != g_id_key and part != g_table_name_key:
                        if part == g_static_var_key:
                            static_vars = data[g_static_var_key]
                            for var in static_vars:
                                new_data[var] = static_vars[var]
                        else:
                            exist_part = data[part]
                            if lang in exist_part:
                                new_data[part] = exist_part[lang]
                            else:
                                new_data[part] = ""
                                logger.warning(f"tofix: table:{table_name} id:{id} part:{part} has no lang:{lang}[{g_code_chinese[lang]}]")
        

        sorted_keys = sorted(normal_map.keys(), key=lambda x: int(x))
        sorted_values = [normal_map[key] for key in sorted_keys]
        doWriteConfigTable(os.path.dirname(file_ab_path),table_name,sorted_values)
    # content = readFile(file_ab_path)
    # if content and len(content) > 0:
    #     origin_content = content
    #     content = doApplyNewDataToDataTxtFile(content,exported_map,lang)
    #     if origin_content != content:
    #         writeFile(file_ab_path,content)

def writeNeedTranslateListToExcel(target_output_path,data_list,same_with_en_no_write=False):
    workbook = xlsxwriter.Workbook(target_output_path)
    sheet = workbook.add_worksheet()
    #第一行表头
    sheet.write('A1','表名')
    sheet.write('B1','id')
    sheet.write('C1','字段名')
    sheet.write('D1',g_code_chinese["en"])
    sheet.write('E1',g_code_chinese["chs"])

    #第二行字段名
    sheet.write('A2','table')
    sheet.write('B2','id')
    sheet.write('C2','key')
    sheet.write('D2',g_en_key)
    sheet.write('E2',g_chs_key)

    #第3行开始数据
    sheet_line_index = 3
    for data in data_list:

        tb = data[g_table_name_key]
        id = data[g_id_key]

        for key in data:
            if key != g_table_name_key and key != g_id_key and key != g_static_var_key and key != g_sort_key:
                all_values = data[key]
                sheet.write('A' + str(sheet_line_index),tb)
                sheet.write('B' + str(sheet_line_index),id)
                sheet.write('C' + str(sheet_line_index),key)

                en_str = ""
                chs_str = ""
                if g_en_key in all_values:
                    en_str = all_values[g_en_key]

                if g_chs_key in all_values:
                    chs_str = all_values[g_chs_key]

                if len(en_str) > 0:
                    sheet.write('D' + str(sheet_line_index),en_str)

                if len(chs_str) > 0 and ( (not same_with_en_no_write) or(same_with_en_no_write and chs_str != en_str) ):
                        sheet.write('E' + str(sheet_line_index),chs_str)

                sheet_line_index = sheet_line_index + 1


    # 保存工作簿到文件
    def saveWorkBookFile():
        try:
            # 保存工作簿到文件
            workbook.close()
            print('save to %s success.' % target_output_path)
        except Exception as e:
            logger.error(f"保存异常: {e}, 3秒后重试")
            time.sleep(3)
            saveWorkBookFile()
            pass
    saveWorkBookFile()

def writeExportBothModifiedListToExcel3(target_output_path,data_list,old_version_path,china_712_path,new_version_path):
    workbook = xlsxwriter.Workbook(target_output_path)
    sheet = workbook.add_worksheet()
    #第一行表头
    sheet.write('A1','表名')
    sheet.write('B1','id')
    sheet.write('C1','字段名')
    sheet.write('D1','原始值')
    sheet.write('E1','810版本值')
    sheet.write('F1','712版本值')

    #第二行字段名
    sheet.write('A2','table')
    sheet.write('B2','id')
    sheet.write('C2','key')
    # sheet.write('D2',g_old_510_key)
    # sheet.write('E2',g_global_810_key)
    # sheet.write('F2',g_china_712_key)
    sheet.write('D2',old_version_path)
    sheet.write('E2',new_version_path)
    sheet.write('F2',china_712_path)

    #第3行开始数据
    sheet_line_index = 3
    for data in data_list:

        tb = data[g_table_name_key]
        id = data[g_id_key]

        for key in data:
            if key != g_table_name_key and key != g_id_key and key != g_static_var_key and key != g_sort_key:
                all_values = data[key]
                sheet.write('A' + str(sheet_line_index),tb)
                sheet.write('B' + str(sheet_line_index),id)
                sheet.write('C' + str(sheet_line_index),key)
                sheet.write('D' + str(sheet_line_index),all_values[g_old_510_key])
                sheet.write("E" + str(sheet_line_index),all_values[g_global_810_key])
                sheet.write("F" + str(sheet_line_index),all_values[g_china_712_key])

                sheet_line_index = sheet_line_index + 1


    # 保存工作簿到文件
    def saveWorkBookFile():
        try:
            # 保存工作簿到文件
            workbook.close()
            print('save to %s success.' % target_output_path)
        except Exception as e:
            logger.error(f"保存异常: {e}, 3秒后重试")
            time.sleep(3)
            saveWorkBookFile()
            pass
    saveWorkBookFile()

def writeExportBothModifiedListToExcel2(target_output_path,data_list,key1_str,key2_str):
    workbook = xlsxwriter.Workbook(target_output_path)
    sheet = workbook.add_worksheet()
    #第一行表头
    sheet.write('A1','表名')
    sheet.write('B1','id')
    sheet.write('C1','字段名')
    sheet.write('D1',key1_str)
    sheet.write('E1',key2_str)

    #第二行字段名
    sheet.write('A2','table')
    sheet.write('B2','id')
    sheet.write('C2','key')
    sheet.write('D2',g_old_key)
    sheet.write('E2',g_new_key)

    #第3行开始数据
    sheet_line_index = 3
    for data in data_list:

        tb = data[g_table_name_key]
        id = data[g_id_key]

        for key in data:
            if key != g_table_name_key and key != g_id_key and key != g_static_var_key and key != g_sort_key:
                all_values = data[key]
                sheet.write('A' + str(sheet_line_index),tb)
                sheet.write('B' + str(sheet_line_index),id)
                sheet.write('C' + str(sheet_line_index),key)
                sheet.write('D' + str(sheet_line_index),all_values[g_old_key])
                sheet.write("E" + str(sheet_line_index),all_values[g_new_key])

                sheet_line_index = sheet_line_index + 1


    # 保存工作簿到文件
    def saveWorkBookFile():
        try:
            # 保存工作簿到文件
            workbook.close()
            print('save to %s success.' % target_output_path)
        except Exception as e:
            logger.error(f"保存异常: {e}, 3秒后重试")
            time.sleep(3)
            saveWorkBookFile()
            pass
    saveWorkBookFile()

# same_with_en_no_write 和 英文文本相同时是否略过不写,默认False全写
def writeToTranslateExcel(data_list,target_output_path):
    workbook = xlsxwriter.Workbook(target_output_path)
    sheet = workbook.add_worksheet()
    #第一行表头
    sheet.write('A1','表名')
    sheet.write('B1','id')
    sheet.write('C1','字段名')
    sheet.write('D1',g_code_chinese["en"])
    sheet.write('E1',g_code_chinese["chs"])

    #第二行字段名
    sheet.write('A2','table')
    sheet.write('B2','id')
    sheet.write('C2','key')
    sheet.write('D2','en')
    sheet.write('E2','chs')
    temp_list = list(g_read_langs)
    temp_list.remove("en")
    temp_list.remove("chs")
    lang_list = sorted(temp_list)


    index = 1
    for l in lang_list:
        sheet.write(get_column_serial("E",index) + '1',g_code_chinese[l])
        sheet.write(get_column_serial("E",index) + '2',l)
        index = index + 1


    #第3行开始数据
    sheet_line_index = 3
    id_index = 1
    for data in data_list:

        id = data[g_id_key]
        tb = data[g_table_name_key]
        for key in data:
            if key != g_table_name_key and key != g_id_key:
                # 对于每一个翻译字段行，如果已经存在中文和英文重复项，则忽略
                value = data[key]
                if data and value:
                    chs_str = ""
                    en_str = ""
                    if g_chs_key in value:
                        chs_str = value[g_chs_key]
                    if g_en_key in value:
                        en_str = value[g_en_key]

                sheet.write('A' + str(sheet_line_index),tb)
                sheet.write('B' + str(sheet_line_index),id)
                sheet.write('C' + str(sheet_line_index),key)
                en_str = ""
                if g_en_key in value:
                    en_str = value[g_en_key]

                if g_chs_key in value:
                    chs_str = value[g_chs_key]

                if len(en_str) > 0:
                    sheet.write('D' + str(sheet_line_index),en_str)

                if len(chs_str) > 0:
                        sheet.write('E' + str(sheet_line_index),chs_str)
                
                # 是否打印其它语言
                index = 1
                for l in lang_list:
                    if l in value:
                        sheet.write(get_column_serial("E",index) + str(sheet_line_index),value[l])
                    index = index + 1
                
                sheet_line_index = sheet_line_index + 1
        id_index = id_index + 1

    # 保存工作簿到文件
    def saveWorkBookFile():
        try:
            # 保存工作簿到文件
            workbook.close()
            print('save to %s success.' % target_output_path)
        except Exception as e:
            logger.error(f"保存异常: {e}, 3秒后重试")
            time.sleep(3)
            saveWorkBookFile()
            pass
    saveWorkBookFile()

def doCompareThreeVersions():
    '''从excel中读取翻译好的数据'''
    new_810_version_map = None
    new_810_version_list = None
    if False:
        # 从 712 目录读取翻译表
        new_810_version_map,new_810_version_list = doWorkReadTranslatedExcels(g_china_712_path)
        # 折叠展开
        decompressMaps(new_810_version_map,new_810_version_list,g_dup_map_from_origin)
        # 展开后追加了很多元素，甚至有重复的，需要重新收集和排序
        new_810_version_list = regerenateVersionList(new_810_version_map)
        new_810_version_list = sortVersionList(new_810_version_list)
    else:
        new_810_version_map,new_810_version_list = doCollectAllLanguagesFromElementDataPath(g_new_version_path)

    # g_china_712_path g_old_version_path 是 datapool\ElementData 目录
    china_712_version_map,china_712_version_list = doCollectAllLanguagesFromElementDataPath(g_china_712_path)
    old_510_version_map,old_510_version_list = doCollectAllLanguagesFromElementDataPath(g_old_version_path)

    # 存储全量的 txt
    saveToTxt(g_china_712_path + os.sep + "new_810_version_list.txt",new_810_version_list)
    saveToTxt(g_china_712_path + os.sep + "china_712_version_list.txt",china_712_version_list)
    saveToTxt(g_china_712_path + os.sep + "old_510_version_list.txt",old_510_version_list)

    # 比校差异，并输出
    doCompareThreeVersionDetails(old_510_version_map,china_712_version_map,new_810_version_map,g_china_712_path,g_old_version_path,g_china_712_path,g_new_version_path)
    pass

def doCompareTwoVersions():
    new_810_version_map,new_810_version_list = doCollectAllLanguagesFromElementDataPath(g_new_version_path)
    old_510_version_map,old_510_version_list = doCollectAllLanguagesFromElementDataPath(g_old_version_path)
    doCompareTwoVersionDetails(old_510_version_map,new_810_version_map,g_new_version_path)

# 把中文翻译表并入英文母表
def doMergeChsToEnBase():
    global g_no_lang_mode
    g_no_lang_mode = False

    path = g_china_712_path
    mother_en_path = os.path.abspath(path + os.sep + "BaseData")
    chs_lang_path = os.path.abspath(path + os.sep +  "/../Localization/ElementData/Loc/CHS")

    collect_chs_list = []
    collect_chs_map = {}
    # 使用收集多语言表的方式收集中文
    doCollectFromMotherDatas(ClassifyFile.LangTables,chs_lang_path,g_chs_key,collect_chs_list,None,collect_chs_map)

    # 使用收集普通表的方式收集EN多语言表
    g_no_lang_mode = True
    collect_en_list = []
    collect_en_map = {}
    doCollectFromMotherDatas(ClassifyFile.LangTables,mother_en_path,g_en_key,collect_en_list,None,collect_en_map)

    # 比较和回写
    # 涉及的表格
    refer_table_names = {}
    # 需要手动操作的部分，即两边都是中文且不同
    manual_deal_results = []
    doMergeChsKeysTo(collect_chs_map,collect_en_map,refer_table_names,manual_deal_results)

    # 回写表格
    writeBackNormalTableToVersion(g_china_712_path,refer_table_names,refer_table_names[g_contain_tables_key])
    export_manual_deal_list = sortVersionList(manual_deal_results)
    if len(manual_deal_results) > 0:
        writeExportBothModifiedListToExcel2(path  + os.sep + "manual_chs.xlsx",export_manual_deal_list,'原始英文表值','中文表值')
    else:
        logger.info(f"no manual deal datas to write to manual_chs.xlsx.")
    pass

# 使用 --new 参数，检查格式化参数是否一致，不一致的输出，多个%d或多个%s的，要警告
def doCheckFormatSerial():
    global g_no_lang_mode
    g_no_lang_mode = False

    path = g_new_version_path
    mother_en_path = os.path.abspath(path + os.sep + "BaseData")
    other_lang_path = os.path.abspath(path + os.sep +  "/../Localization/ElementData/Loc")

    collect_list = []
    collect_map = {}
    doWorkCollectOnFolderAndExport(ClassifyFile.LangTables,mother_en_path,other_lang_path,collect_list,None,collect_map,True)
    doCollectErrorFormatAndWarningFormat(path,collect_map,collect_list)

# 修复路径下的所有引号问题，仅限多语言文件
def fixDoubleQuote(replace_char):
    global g_no_lang_mode
    g_no_lang_mode = False

    path = g_new_version_path
    mother_en_path = os.path.abspath(path + os.sep + "BaseData")
    other_lang_path = os.path.abspath(path + os.sep +  "/../Localization/ElementData/Loc")
    if mother_en_path and len(mother_en_path) > 0:
        for root, _, files in os.walk(mother_en_path):
            for file in files:
                file_ab_path = os.path.join(root, file)
                if file.endswith("_LANGUAGE.data.txt"):
                    fix_double_quote(file_ab_path,'utf-16',replace_char)

    files_and_folders = os.listdir(other_lang_path)
    for item in files_and_folders:
        dir_ab_path = os.path.join(other_lang_path, item)
        if os.path.isdir(dir_ab_path):
            lang = item.lower()
            if lang not in g_read_langs:
                g_read_langs[lang] = lang
            for root, _, files in os.walk(dir_ab_path):
                for file in files:
                    file_ab_path = os.path.join(root, file)
                    if file.endswith("_LANGUAGE.data.txt"):
                        fix_double_quote(file_ab_path,'utf-16',replace_char)

# 
def doCollectErrorFormatAndWarningFormat(path,collect_map,collect_list):
    error_list = []
    warning_list = []
    for key_pair in collect_map:
        table_info = collect_map[key_pair]
        if g_table_name_key in table_info and g_id_key in table_info and len(table_info) > 2:
            for part in table_info:
                if part != g_table_name_key and part != g_id_key and part != g_static_var_key:
                    # 普通表时
                    lang_map = table_info[part]
                    doCheckErrorAndWarningFormat(g_chs_key,table_info[g_table_name_key],table_info[g_id_key],part,lang_map,error_list,warning_list)

    doWriteFormatListToExcel(path  + os.sep + "format_error_list.xlsx",error_list,"中文","翻译语言",g_chs_key)
    doWriteFormatListToExcel(path  + os.sep + "format_warning_list.xlsx",warning_list,"中文","翻译语言",g_chs_key)


def doWriteFormatListToExcel(target_output_path,data_list,key1_str,key2_str,main_key):
    workbook = xlsxwriter.Workbook(target_output_path)
    sheet = workbook.add_worksheet()
    #第一行表头
    sheet.write('A1','表名')
    sheet.write('B1','id')
    sheet.write('C1','字段名')
    sheet.write('D1',key1_str)
    sheet.write('E1','语言')
    sheet.write('F1',key2_str)

    #第二行字段名
    sheet.write('A2','table')
    sheet.write('B2','id')
    sheet.write('C2','key')
    sheet.write('D2',main_key)
    sheet.write('E2','lang')
    sheet.write('F2',g_content_key)

    #第3行开始数据
    sheet_line_index = 3
    for data in data_list:

        tb = data[g_table_name_key]
        id = data[g_id_key]

        for key in data:
            if key != g_table_name_key and key != g_id_key and key != g_static_var_key and key != g_sort_key:
                all_values = data[key]
                sheet.write('A' + str(sheet_line_index),tb)
                sheet.write('B' + str(sheet_line_index),id)
                sheet.write('C' + str(sheet_line_index),key)
                sheet.write('D' + str(sheet_line_index),all_values[main_key])
                other_lang_i = 0
                for other_lang in all_values:
                    if other_lang != main_key:
                        other_lang_i = other_lang_i + 1
                        sheet.write(get_column_serial("D",other_lang_i) + str(sheet_line_index),other_lang)
                        other_lang_i = other_lang_i + 1
                        sheet.write(get_column_serial("D",other_lang_i) + str(sheet_line_index),all_values[other_lang])

                sheet_line_index = sheet_line_index + 1


    # 保存工作簿到文件
    def saveWorkBookFile():
        try:
            # 保存工作簿到文件
            workbook.close()
            print('save to %s success.' % target_output_path)
        except Exception as e:
            logger.error(f"保存异常: {e}, 3秒后重试")
            time.sleep(3)
            saveWorkBookFile()
            pass
    saveWorkBookFile()


def doCheckErrorAndWarningFormat(base_key,table_name,table_id,part,lang_map,error_list,warning_list):
    var_base_str = None
    if base_key == g_chs_key and base_key not in lang_map:
        # 中文里没有时，取用英文
        logger.warning(f"doCheckErrorAndWarningFormat {table_name} {table_id} {part} has no {base_key} lang." )
        base_key = g_en_key
    var_base_str = lang_map[base_key]
    if var_base_str.find("%") < 0:
        return
    format_map = _calc_str_format_map(var_base_str)
    if len(format_map) > 0:
        for lang_key in lang_map:
            if lang_key != base_key:
                l_str = lang_map[lang_key]
                l_format_map = _calc_str_format_map(l_str)
                if g_count_key not in l_format_map or format_map[g_count_key] != l_format_map[g_count_key] or format_map[g_serial_key] != l_format_map[g_serial_key]:
                    # 数量不对，一定有问题
                    # 顺序不一致，一定有问题
                    error_list.append({g_table_name_key:table_name,g_id_key:table_id,part:{g_chs_key:var_base_str,lang_key:l_str}})
                else:
                    if format_map[g_count_key] > 1 and format_map[g_all_same_key]:
                        warning_list.append({g_table_name_key:table_name,g_id_key:table_id,part:{base_key:var_base_str,lang_key:l_str}})
    pass


# collect_chs_map 收集好的中文表
# collect_en_map 收集好的英文表
# refer_table_names 涉及的所有表名，用于回写
# manual_deal_results 需要手动确认的部分
def doMergeChsKeysTo(collect_chs_map,collect_en_map,refer_table_names,manual_deal_results):
    table_name_map = {}
    refer_table_names[g_contain_tables_key] = table_name_map
    for tb_id in collect_chs_map:
        tb_data = collect_chs_map[tb_id]
        if g_table_name_key in tb_data:
            table_name = tb_data[g_table_name_key]
            table_id = tb_data[g_id_key]
            for part in tb_data:
                if part != g_table_name_key and part != g_id_key and part != g_static_var_key:
                    en_table_datas = collect_en_map[table_name]
                    if table_id in en_table_datas:
                        en_table_data = en_table_datas[table_id]
                        if part in en_table_data:
                            en_value = en_table_data[part]
                            chs_value = tb_data[part][g_chs_key]
                            if has_chinese(en_value):
                                if en_value != chs_value:
                                    manual_deal_results.append({g_table_name_key:table_name,g_id_key:table_id,g_sort_key:part,part:{g_old_key:en_value,g_new_key:chs_value}})
                                    # 双方都是中文时，也回写
                                    #######################################
                                    en_table_data[part] = chs_value

                                    if not table_name in table_name_map:
                                        table_name_map[table_name] = collect_en_map[table_name]
                                    #######################################
                            else:
                                #######################################
                                en_table_data[part] = chs_value

                                if not table_name in table_name_map:
                                    table_name_map[table_name] = collect_en_map[table_name]
                                #######################################
                        else:
                            logger.warning(f"BaseData table:{table_name} id:{table_id} has no part:{part}, which chs has.")

# 收集所有配置的id序列，得到一个基础表
def doCollectAllTableId():
    global g_no_lang_mode
    g_no_lang_mode = True

    path = g_new_version_path
    mother_en_path = os.path.abspath(path + os.sep + "BaseData")
    templ_id_dat_file = os.path.abspath(path + os.sep + "TemplID.dat")
    other_lang_path = os.path.abspath(path + os.sep +  "/../Localization/ElementData/Loc")
    collect_list = []
    collect_map = {}
    # id 用来记录重复项
    # tb 用来记录重复的表名
    collect_dup_map = {g_id_key:{},g_table_name_key:{},g_dup_key:[]}
    withTranslated = False
    table_classify = ClassifyFile.AllTables
    doWorkCollectOnFolderAndExport(table_classify,mother_en_path,other_lang_path,collect_list,collect_dup_map,collect_map,withTranslated)
    total_count = len(collect_dup_map[g_id_key])
    logger.info(f"total count:{total_count}")
    dup_arr = collect_dup_map[g_dup_key]
    logger.info(f"dup count:{len(dup_arr)}")
    if len(dup_arr) > 0:
        for dup_info_str in dup_arr:
            logger.info(dup_info_str)

        writeFile(path + os.sep + "exist_dup_list.txt","\n".join(dup_arr) + f"\ndup count:{len(dup_arr)}")

    with open(path + os.sep + "origin_all_table_id.txt", 'wb') as file:
        pickle.dump(collect_dup_map, file)

# 修复整个配置表中多出的重复id，自动生成新值，并将最大值回写进 TemplID.dat
def doFixDuplicateTableId():
    global g_no_lang_mode
    g_no_lang_mode = True

    path = g_new_version_path
    mother_en_path = os.path.abspath(path + os.sep + "BaseData")
    templ_id_dat_file = os.path.abspath(path + os.sep + "TemplID.dat")
    other_lang_path = os.path.abspath(path + os.sep +  "/../Localization/ElementData/Loc")
    # if mother_en_path and len(mother_en_path) > 0:
    #     for root, _, files in os.walk(mother_en_path):
    #         for file in files:
    #             file_ab_path = os.path.join(root, file)
    #             if file.endswith("_LANGUAGE.data.txt"):
    #                 pass

    # files_and_folders = os.listdir(other_lang_path)
    # for item in files_and_folders:
    #     dir_ab_path = os.path.join(other_lang_path, item)
    #     if os.path.isdir(dir_ab_path):
    #         lang = item.lower()
    #         if lang not in g_read_langs:
    #             g_read_langs[lang] = lang
    #         for root, _, files in os.walk(dir_ab_path):
    #             for file in files:
    #                 file_ab_path = os.path.join(root, file)
    #                 if file.endswith("_LANGUAGE.data.txt"):
    #                     pass
    collect_list = []
    collect_map = {}
    # id 用来记录重复项
    # tb 用来记录重复的表名
    collect_dup_map = {g_id_key:{},g_table_name_key:{}}
    withTranslated = True
    table_classify = ClassifyFile.AllTables
    doWorkCollectOnFolderAndExport(table_classify,mother_en_path,other_lang_path,collect_list,collect_dup_map,collect_map,withTranslated)

def collect_line_endings(file_path):
    with open(file_path, 'rb') as file:
        # 移动到文件末尾
        file.seek(0, 2)
        file_size = file.tell()
        
        if file_size == 0:
            return []  # 文件为空
        
        # 向前读取一定的字节
        file.seek(max(0, file_size - 20), 0)  # 读取最后20个字节（可调整）
        content = file.read()

        # 收集连续的 \r, \n 和 \0 字符
        sequences = ""
        current_sequence = []

        # 倒序遍历
        for byte in reversed(content):
            if byte in (b'\r'[0], b'\n'[0], b'\0'[0]):
                # current_sequence.append('\\r' if byte == b'\r'[0] else '\\n' if byte == b'\n'[0] else '\\0')
                 if byte == b'\r'[0] or byte == b'\n'[0]:
                    current_sequence.append(chr(byte))
            else:
                break

        # 处理最后一个序列
        if current_sequence:
            # if len(current_sequence) % 2 == 1:
            #     current_sequence.pop()
            sequences = ''.join(reversed(current_sequence))

        return sequences

def doTestReplace(file_path):
    # 如果之前的文件后部有很多空白，保留之
    origin_content = readFile(file_path)
    # empty_chars = collect_line_endings(file_path)
    # 回写，和原来的格式一模一样
    writeFile(file_path,origin_content + "\n")


# 把多语言表中缺失的字段,用母表中的字段填充
def doLangPartFillToFull():
    # test
    # doTestReplace("D:\\R\\datapool\\ElementData\\BaseData\\ROD_SKIN_LANGUAGE.data.txt")
    # return
    # end
    global g_no_lang_mode
    g_no_lang_mode = False

    global g_new_version_path
    if len(g_new_version_path) <= 0:
        file_path = os.path.abspath(__file__)
        g_new_version_path = os.path.dirname(file_path)

    path = g_new_version_path
    mother_en_path = os.path.abspath(path + os.sep + "BaseData")
    other_lang_path = os.path.abspath(path + os.sep +  "/../Localization/ElementData/Loc")
    collect_list = []
    collect_map = {}
    collect_dup_map = None
    withTranslated = True
    table_classify = ClassifyFile.LangTables
    doWorkCollectOnFolderAndExport(table_classify,mother_en_path,other_lang_path,collect_list,collect_dup_map,collect_map,withTranslated)
    converted_table_map = {}
    for k in collect_map:
        info = collect_map[k]
        if g_table_name_key in info:
            table_name = info[g_table_name_key]
            id = info[g_id_key]
            # tables_map = None
            # if table_name not in converted_table_map:
            #     tables_map = {}
            #     converted_table_map[table_name] = tables_map
            # else:
            #     tables_map = converted_table_map[table_name]
            # tables_map[id] = info
            converted_table_map[k] = info

            for part in info:
                if part != g_table_name_key and part != g_id_key and part != g_static_var_key:
                    part_info = info[part]
                    for lang in g_read_langs:
                        if lang != g_en_key and g_en_key in part_info and lang not in part_info:
                            part_info[lang] = part_info[g_en_key]
                            logger.info(f"do add part for {table_name} {id} {part} fill en to {lang}, content:{part_info[g_en_key]}")


    convertLangWriteBackMap(converted_table_map)
    # 原理不同，语言表是替换翻译内容的方式
    writeBackLangToVersion(path,converted_table_map,True)
    pass

def doCollectAnalaysedError(table_name,id,part,part_info,error_list):
    each_count = -1
    first_count_lang = ""
    first_lang_str = ""

    if g_chs_key in part_info:
        first_lang_str = part_info[g_chs_key]
        first_count_lang = g_chs_key
        each_count = first_lang_str.count('%')

    elif g_en_key in part_info:
        first_lang_str = part_info[g_en_key]
        first_count_lang = g_en_key
        each_count = first_lang_str.count('%')

    for lang in part_info:
        lang_str = part_info[lang]
        f_char_count = lang_str.count('%')
        if each_count == -1:
            each_count = f_char_count
            first_count_lang = lang
            first_lang_str = lang_str
        else:
            if each_count != f_char_count:
                info = {g_table_name_key:table_name,g_id_key:id,g_part_key:part,g_value_key:part_info,part:f'{table_name} {id} {part} %count {lang}:{lang_str} count:{f_char_count} not match {first_count_lang}:{first_lang_str} count:{each_count}'}
                info["excel_data"] = {g_table_name_key:table_name,g_id_key:id,part:{first_count_lang:first_lang_str,lang:lang_str}}
                error_list.append(info)


def doScanTranslate():
    global g_no_lang_mode
    g_no_lang_mode = False

    global g_new_version_path
    if len(g_new_version_path) <= 0:
        file_path = os.path.abspath(__file__)
        g_new_version_path = os.path.dirname(file_path)

    path = g_new_version_path
    mother_en_path = os.path.abspath(path + os.sep + "BaseData")
    other_lang_path = os.path.abspath(path + os.sep +  "/../Localization/ElementData/Loc")
    collect_list = []
    collect_map = {}
    collect_dup_map = None
    withTranslated = True
    table_classify = ClassifyFile.LangTables
    doWorkCollectOnFolderAndExport(table_classify,mother_en_path,other_lang_path,collect_list,collect_dup_map,collect_map,withTranslated)
    # converted_table_map = {}
    error_list = []
    for k in collect_map:
        info = collect_map[k]
        if g_table_name_key in info:
            table_name = info[g_table_name_key]
            id = info[g_id_key]
            # tables_map = None
            # if table_name not in converted_table_map:
            #     tables_map = {}
            #     converted_table_map[table_name] = tables_map
            # else:
            #     tables_map = converted_table_map[table_name]
            # tables_map[id] = info
            # converted_table_map[k] = info

            for part in info:
                if part != g_table_name_key and part != g_id_key and part != g_static_var_key:
                    part_info = info[part]
                    doCollectAnalaysedError(table_name,id,part,part_info,error_list)

    excel_data_list = []
    for error_data in error_list:
        logger.error(f"{error_data[g_table_name_key]} {error_data[g_id_key]} {error_data[g_part_key]} msg:{error_data[error_data[g_part_key]]}")
        excel_data_list.append(error_data["excel_data"])

    logger.error(f"total count:{len(error_list)}")

    writeToTranslateExcel(excel_data_list,path  + os.sep + "str_format_error_list.xlsx")

    # convertLangWriteBackMap(converted_table_map)
    # # 原理不同，语言表是替换翻译内容的方式
    # writeBackLangToVersion(path,converted_table_map,True)

def doFixIdToTemplateId():
    global g_no_lang_mode
    g_no_lang_mode = False

    global g_new_version_path
    if len(g_new_version_path) <= 0:
        file_path = os.path.abspath(__file__)
        g_new_version_path = os.path.dirname(file_path)

    path = g_new_version_path
    mother_en_path = os.path.abspath(path + os.sep + "BaseData")
    other_lang_path = os.path.abspath(path + os.sep +  "/../Localization/ElementData/Loc")

    if len(g_op_files) > 0:
            arr = g_op_files.split('|')
            if len(arr) > 0:
                for file_path in arr:
                    collect_table_names_map = {}
                    readTranslatedExcel(file_path,collect_table_names_map,None)

                    collect_list = []
                    collect_map = {}
                    collect_dup_map = None
                    withTranslated = True
                    table_classify = ClassifyFile.LangTables
                    doWorkCollectOnFolderAndExport(table_classify,mother_en_path,other_lang_path,collect_list,collect_dup_map,collect_map,withTranslated,collect_table_names_map)

                    template_names = {}

                    converted_table_map = {}
                    for k in collect_map:
                        info = collect_map[k]
                        if g_table_name_key in info:
                            table_name = info[g_table_name_key]
                            id = info[g_id_key]

                            static_value = info[g_static_var_key]
                            template_value = None
                            template_name = None
                            if table_name in template_names:
                                template_name = template_names[table_name]
                            if not template_name:
                                for k in static_value:
                                    if k != g_id_key and k != g_name_key:
                                        template_name = k
                                        break
                            if template_name:
                                template_value = static_value[template_name]
                                if not table_name in template_names:
                                    template_names[table_name] = template_name
                            
                            if template_value and template_value != "" and template_value != "0":
                                # 关键操作，覆盖 id
                                info[g_id_key] = template_value
                                static_value[g_id_key] = template_value

                            converted_table_map[f"{table_name}^{template_value or id}"] = info

                            for part in info:
                                if part != g_table_name_key and part != g_id_key and part != g_static_var_key:
                                    part_info = info[part]
                                    for lang in g_read_langs:
                                        if lang != g_en_key and g_en_key in part_info and lang not in part_info:
                                            part_info[lang] = part_info[g_en_key]
                                            logger.info(f"do add part for {table_name} {id} {part} fill en to {lang}, content:{part_info[g_en_key]}")


                    convertLangWriteBackMap(converted_table_map)
                    # 原理不同，语言表是替换翻译内容的方式
                    writeBackLangToVersion(path,converted_table_map,True,True)

                    # for table_name in collect_table_names_map:
                    #     file_ab_path = mother_en_path + os.sep + table_name.upper() + ".data.txt"
                    #     collect_map = {}
                    #     collect_list = []
                    #     template_name = None
                    #     doWorkOnDataTxtFile(file_ab_path,g_en_key,collect_list,None,collect_map)
                    #     for table_info in collect_list:
                    #         static_value = table_info[g_static_var_key]
                    #         template_value = None
                    #         if not template_name:
                    #             for k in static_value:
                    #                 if k != g_id_key and k != g_name_key:
                    #                     template_name = k
                    #                     break
                    #         if template_name:
                    #             template_value = static_value[template_name]
                    #         if template_value and template_value != "" and template_value != "0":
                    #             # 关键操作，覆盖 id
                    #             table_info[g_id_key] = template_value
                    #             static_value[g_id_key] = template_value
                        
                    #     doWriteConfigTable(os.path.dirname(file_ab_path),table_name,collect_list)

                    #    pass
                    pass
    pass

def usage():
    print("translate_assist_collect.py: collect languages to translate, write to xlsx files.\n"
          "args: \n"
          " --in path   eg: .\BaseData\n"
          " --op op  eg: writeback(回写进语言数据表) default is collect\n"
          "eg: \n"
          " translate_assist_collect.py -in .\BaseData --op writeback")
    
def get_opt():
    try:
        options, args = getopt.getopt(sys.argv[1:], "", ["in=", "op=","files=", "help","old=","new=","china=","notlangtable="])
        for name, value in options:
            # if name == '--in':
            #     global g_china_712_path
            #     g_china_712_path = os.path.abspath(value + "/../")
            #     # global g_other_lang_path
            #     # g_other_lang_path = os.path.abspath(g_china_712_path + "/../Localization/ElementData/Loc")
            # el
            if name == '--op':
                global g_op_str
                g_op_str = value
            if name == '--files':
                global g_op_files
                g_op_files = value
            elif name == '--old':
                global g_old_version_path
                g_old_version_path = value
            elif name == '--new':
                global g_new_version_path
                g_new_version_path = value
            elif name == '--china':
                global g_china_712_path
                g_china_712_path = value
            elif name == '--notlangtable':
                global g_no_lang_mode
                g_no_lang_mode = int(value) == 1
            elif name == '--help':
                usage()
                return False

        if g_op_str in g_op_enum_map:
            global g_op
            g_op = g_op_enum_map[g_op_str]
        else:
            usage()
            return False
        return True
    except Exception as e:
        print("get_opt error: %s" % (e))
        usage()
        return False

def main(argv):
    print(len(argv),argv)
    file_path = os.path.abspath(__file__)
    dir_path = os.path.dirname(file_path)

    global g_dup_map_from_origin
    # origin_dup_list_path = "D:\\work\\fishing\\svn\\Dev_Global_2024_5_10_200M\\datapool\\ElementData\\origin_dup_list.txt"
    origin_dup_list_path = "./origin_dup_list.txt"
    if os.path.exists(origin_dup_list_path):
        with open(origin_dup_list_path, 'rb') as file:
            # g_dup_map_from_origin = pickle.load(file)
            pass

    if get_opt():
        if g_op == Op.Collect or g_op == Op.CollectAll:
            logger.info("please run another tool.")
        elif g_op == Op.WriteBack:
            logger.info("please run another tool.")
        elif g_op == Op.GenerateDiffPart:
            logger.info("please run another tool.")
        elif g_op == Op.Generate3VersionDiff:
            doCompareThreeVersions()
            # doTest()
            # doTestDumpData()
        elif g_op == Op.Generate2VersionDiff:
            doCompareTwoVersions()
        elif g_op == Op.MergeChsToEnBase:
            doMergeChsToEnBase()
        elif g_op == Op.CheckFormatSerial:
            doCheckFormatSerial()
        elif g_op == Op.DoubleQuote:
            # 把所有多语言.data.txt中双引号内的双引号全部替换为指定字符
            # 这一行填''代表空字符即直接删除
            fixDoubleQuote(g_double_quote_replace_char)
        elif g_op == Op.CollectAllTableId:
            # 收集所有配置的id序列，得到一个基础表
            doCollectAllTableId()
        elif g_op == Op.FixDuplicateTableId:
            # 修复整个配置表中多出的重复id，自动生成新值，并将最大值回写进 TemplID.dat
            doFixDuplicateTableId()
        elif g_op == Op.LangPartFillToFull:
            doLangPartFillToFull()
        elif g_op == Op.ScanTranslate:
            doScanTranslate()
        elif g_op == Op.FixIdToTemplateId:
            doFixIdToTemplateId()
        else:
            logger.info("nothing to do.")
    

if __name__ == "__main__":
    argv_len = len(sys.argv)
    start_time = time.time()
    main(sys.argv)

    end_time = time.time()
    total_seconds = end_time - start_time
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    logger.info(f"done in {int(hours)}时 {int(minutes)}分 {seconds:.3f}秒")
    if len(all_need_collector_log) > 0:
        print("all errors summary:")
        for log in all_need_collector_log:
            print("\t",log)

    else:
        print("no error logs.")


