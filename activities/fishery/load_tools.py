import ast
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from tools.excelRead import ExcelToolsForActivities


def get_worksheet():
    path = "tools.xlsx"
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook["Sheet1"]
    return worksheet

def get_row_color_bg(worksheet: Worksheet, column_start, row_index, table_data_len):
    row_color_bg = []
    cur = column_start
    while cur <= table_data_len + column_start:
        row_color_bg.append(worksheet.cell(row_index, cur).fill.fgColor.rgb)
        cur += 1
    return row_color_bg


def get_row_color_font(worksheet: Worksheet, column_start, row_index, table_data_len):
    row_color_font = []
    cur = column_start
    while cur <= table_data_len + column_start:
        cell = worksheet.cell(row_index, cur)
        if cell.value is None:
            row_color_font.append("")
            cur += 1
            continue
        font_color = cell.font.color.rgb
        if type(font_color) is str:
            row_color_font.append(font_color)
        else:
            row_color_font.append("")
        cur += 1
    return row_color_font

def get_row_data(worksheet, row_index, column_start, table_data_len):
    row_data = []
    # 第六行开始
    cur = column_start
    while cur <= table_data_len + column_start:
        row_data.append(worksheet.cell(row_index, cur).value)
        cur += 1
    return row_data

def get_column_data(worksheet, column_index, row_start):
    column_data = []
    cur = row_start
    data = worksheet.cell(cur, column_index).value
    while data:
        column_data.append(data)
        cur += 1
        data = worksheet.cell(cur, column_index).value
    return column_data

def get_exclude_info(excel_tool: ExcelToolsForActivities, fishery_id):
    fish_list = excel_tool.get_fish_id_list(fishery_id=fishery_id)
    worksheet = get_worksheet()
    color_bg_list = []
    color_font_list = []
    cur = 8
    while cur < 16:
        color_bg_list.append(get_row_color_bg(worksheet, row_index=cur, column_start=8, table_data_len=15))
        color_font_list.append(get_row_color_font(worksheet, row_index=cur, column_start=8, table_data_len=15))
        cur += 1
    exclude_info = []
    cur = 0
    while cur < 8:
        info = []
        i = 0
        while i < len(color_bg_list[cur]):
            color = color_bg_list[cur][i]
            if color == "FFA9D08D":
                info.append(fish_list[i + 15])
                i += 1
                continue
            if color_font_list[cur][i] == "FFFE0300":
                info.append(fish_list[i + 15])
                i += 1
                continue
            i += 1
        exclude_info.append(info)
        cur += 1
    return exclude_info

def get_fishery_info(excel_tool: ExcelToolsForActivities, fishery_id):
    worksheet = get_worksheet()
    fishery_info = []
    fish_id_list = excel_tool.get_fish_id_list(fishery_id=fishery_id)

    fish_type_info = get_row_data(worksheet, row_index=6, column_start=8, table_data_len=15)
    fish_spot_info_common = []
    fish_spot_info_rare = []

    cur = 8
    while cur < 16:
        fish_spot_info_rare.append(get_row_data(worksheet, row_index=cur, column_start=8, table_data_len=15))
        cur += 1

    cur = 21
    while cur < 29:
        fish_spot_info_common.append(get_row_data(worksheet, row_index=cur, column_start=8, table_data_len=15))
        cur += 1

    fish_class_info = get_row_data(worksheet, row_index=16, column_start=8, table_data_len=15)
    fish_battle_type_info = get_row_data(worksheet, row_index=17,column_start=8, table_data_len=15)
    fish_ai_info_common = get_row_data(worksheet, row_index=29, column_start=8, table_data_len=15)
    fish_ai_info_rare = get_row_data(worksheet, row_index=30, column_start=8, table_data_len=15)

    cur = 0
    while cur < 15:
        fish_info = {}
        fish_info["tpId"] = fish_id_list[cur]
        if fish_type_info[cur] == 'S':
            fish_info["fishType"] = 1
        elif fish_type_info[cur] == 'M':
            fish_info["fishType"] = 2
        elif fish_type_info[cur] == 'L':
            fish_info["fishType"] = 3
        elif fish_type_info[cur] == 'H':
            fish_info["fishType"] = 4
        elif fish_type_info[cur] == 'G':
            fish_info["fishType"] = 5
        fish_info["fishSpot"] = []
        i = 0
        while i < 8:
            if fish_spot_info_common[i][cur] is None:
                i += 1
                continue
            fish_info["fishSpot"].append(i)
            i += 1
        fish_info["fishClass"] = 1
        fish_info["fishAI"] = fish_ai_info_common[cur]
        if fish_battle_type_info[cur] == "力":
            fish_info["newPlotBattleType"] = 1
        if fish_battle_type_info[cur] == "敏":
            fish_info["newPlotBattleType"] = 2
        if fish_battle_type_info[cur] == "智":
            fish_info["newPlotBattleType"] = 3
        fishery_info.append(fish_info)
        cur += 1

    while cur < 30:
        index = cur - 15
        fish_info = {}
        fish_info["tpId"] = fish_id_list[cur]
        fish_info["fishType"] = fishery_info[index]["fishType"]
        fish_info["fishSpot"] = []
        i = 0
        while i < 8:
            if fish_spot_info_rare[i][index] is None:
                i += 1
                continue
            fish_info["fishSpot"].append(i)
            i += 1
        if fish_class_info[index] == 'R':
            fish_info["fishClass"] = 2
        elif fish_class_info[index] == 'E':
            fish_info["fishClass"] = 3
        elif fish_class_info[index] == 'M':
            fish_info["fishClass"] = 4
        fish_info["fishAI"] = fish_ai_info_rare[index]
        if fish_battle_type_info[index] == "力":
            fish_info["newPlotBattleType"] = 1
        if fish_battle_type_info[index] == "敏":
            fish_info["newPlotBattleType"] = 2
        if fish_battle_type_info[index] == "智":
            fish_info["newPlotBattleType"] = 3
        fishery_info.append(fish_info)
        cur += 1
    return fishery_info

def get_spot_fish_type_detail():
    worksheet = get_worksheet()
    spot_fish_type_detail = []
    cur = 3
    while cur < 11:
        fish_type_detail = {"small": 0, "medium": 0, "large": 0, "hidden": 0, "boss": 0, "rare": 0, "elite": 0,
                            "monster": 0, "total_rare": 0, "total_common": 0}
        spot_fish_type_info = get_row_data(worksheet, row_index=cur, column_start=3, table_data_len=4)
        fish_type_detail["total_rare"] = spot_fish_type_info[0]
        fish_type_detail["rare"] = spot_fish_type_info[1]
        fish_type_detail["elite"] = spot_fish_type_info[2]
        fish_type_detail["monster"] = spot_fish_type_info[3]
        spot_fish_type_info = get_row_data(worksheet, row_index=cur + 13, column_start=1, table_data_len=6)
        fish_type_detail["total_common"] = spot_fish_type_info[0]
        fish_type_detail["small"] = spot_fish_type_info[1]
        fish_type_detail["medium"] = spot_fish_type_info[2]
        fish_type_detail["large"] = spot_fish_type_info[3]
        fish_type_detail["hidden"] = spot_fish_type_info[4]
        fish_type_detail["boss"] = spot_fish_type_info[5]
        spot_fish_type_detail.append(fish_type_detail)
        cur += 1
    return spot_fish_type_detail

def get_quest_info(excel_tool: ExcelToolsForActivities, fishery_id):
    fish_list = excel_tool.get_fish_id_list(fishery_id=fishery_id)
    worksheet = get_worksheet()
    color_list = []
    cur = 8
    while cur < 16:
        color_list.append(get_row_color_bg(worksheet, row_index=cur, column_start=8, table_data_len=15))
        cur += 1
    quest_info = []
    cur = 0
    while cur < 8:
        i = 0
        while i < len(color_list[cur]):
            color = color_list[cur][i]
            if color != "FFA9D08D":
                i += 1
                continue
            fish_id = fish_list[i + 15]
            if fish_id in quest_info:
                i += 1
                continue
            quest_info.append(fish_id)
            i += 1
        cur += 1
    return list(quest_info)

def get_cfg_common(column):
    worksheet = get_worksheet()
    column_data = get_column_data(worksheet=worksheet, column_index=column, row_start=46)
    info = {}
    for data in column_data:
        data_split = data.split(" = ")
        key = data_split[0]
        value = ast.literal_eval(data_split[1])
        info[key] = value
    return info


def get_cfg_fish_bag():
    return get_cfg_common(3)


def get_cfg_fishery():
    return get_cfg_common(6)


def get_cfg_achievement():
    return get_cfg_common(9)


def get_cfg_flash_card():
    return get_cfg_common(12)


def get_cfg_ndays():
    return get_cfg_common(15)


if __name__ == '__main__':
    res = get_cfg_ndays()
    print(res)
